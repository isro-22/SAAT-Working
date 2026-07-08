from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill

import src.ui_theme as ui_theme

CUSTOM_CSS = ui_theme.CUSTOM_CSS
render_sheet_status_pill = ui_theme.render_sheet_status_pill


APP_TITLE = "Clean Database"
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DEFAULT_WORKBOOK = DATA_DIR / "Bahan.xlsx"
OUTPUT_SHEET = "BOL-SAAT List"
VALID_SEGMENTS = ["AC", "CAS", "CF", "CHM", "CM", "CS", "FBL", "FL", "NAT", "SV", "SW"]
ERROR_VALUES = {"#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}
CODE_PATTERN = re.compile(rf"^GEN-({'|'.join(VALID_SEGMENTS)})-\d{{5}}$")


st.set_page_config(page_title=APP_TITLE, layout="wide", initial_sidebar_state="expanded")


def clean_text(value: object) -> str:
    """Normalize display text without changing meaningful inner spaces."""
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_code(value: object) -> str:
    """Clean item code for matching and validation."""
    return clean_text(value).upper()


def to_number(value: object) -> float:
    number = pd.to_numeric(value, errors="coerce")
    if pd.isna(number):
        return 0.0
    return float(number)


def pick_column(df: pd.DataFrame, options: list[str], fallback_index: int | None = None) -> str:
    normalized = {clean_text(col).lower(): col for col in df.columns}
    for option in options:
        option = option.lower()
        for key, original in normalized.items():
            if option in key:
                return original
    if fallback_index is not None and fallback_index < len(df.columns):
        return df.columns[fallback_index]
    raise ValueError(f"Column not found: {', '.join(options)}")


@st.cache_data(show_spinner=False)
def read_workbook(file: str | Path | BinaryIO) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    bol = pd.read_excel(file, sheet_name="01 BOL DB", dtype=object)
    saat = pd.read_excel(file, sheet_name="01 SAAT DB", dtype=object)
    bins = pd.read_excel(file, sheet_name="02 SAAT BIN", header=1, dtype=object)
    bol.columns = [clean_text(col) for col in bol.columns]
    saat.columns = [clean_text(col) for col in saat.columns]
    bins.columns = [clean_text(col) for col in bins.columns]
    return bol, saat, bins


@st.cache_data(show_spinner=False)
def read_sheet_names(file: str | Path | BinaryIO) -> list[str]:
    if hasattr(file, "seek"):
        file.seek(0)
    names = pd.ExcelFile(file).sheet_names
    if hasattr(file, "seek"):
        file.seek(0)
    return names


def validate_code_series(series: pd.Series, source: str) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for row_number, raw_value in series.items():
        original = "" if pd.isna(raw_value) else str(raw_value)
        stripped = clean_code(raw_value)

        if pd.isna(raw_value) or stripped == "":
            status = "KOSONG"
        elif stripped in ERROR_VALUES:
            status = "ERROR FORMULA"
        elif CODE_PATTERN.match(stripped):
            status = "VALID (ADA SPASI TERSEMBUNYI)" if original != original.strip() else "VALID"
        else:
            status = "INVALID FORMAT"

        segment = ""
        match = re.match(r"^GEN-([A-Z0-9]+)-", stripped)
        if status.startswith("VALID") and match:
            segment = match.group(1)

        rows.append(
            {
                "Source": source,
                "Excel Row": int(row_number) + 2,
                "Original Code": original,
                "Clean Code": stripped,
                "Segment": segment,
                "Status": status,
            }
        )
    return pd.DataFrame(rows)


def first_lookup(df: pd.DataFrame, key_col: str, value_col: str, clean_key: bool = True) -> dict[str, object]:
    lookup: dict[str, object] = {}
    for _, row in df.iterrows():
        key = clean_code(row[key_col]) if clean_key else clean_text(row[key_col])
        if key and key not in lookup:
            lookup[key] = row[value_col]
    return lookup


def build_output(
    bol: pd.DataFrame,
    saat: pd.DataFrame,
    bins: pd.DataFrame,
    usd_to_idr: float,
    manual_chemicals: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, int]]:
    bol_code_col = pick_column(bol, ["nav code"], fallback_index=11)
    bol_chemical_col = pick_column(bol, ["chemical name"], fallback_index=1)
    bol_item_col = pick_column(bol, ["segment 1+2+3+4"], fallback_index=7)
    bol_price_col = pick_column(bol, ["commercial price", "control price"], fallback_index=10)
    bol_cas_col = pick_column(bol, ["cas no"], fallback_index=9)
    bol_appearance_col = pick_column(bol, ["physical state"], fallback_index=22)

    saat_code_col = pick_column(saat, ["no."], fallback_index=0)
    saat_item_col = pick_column(saat, ["description"], fallback_index=1)
    saat_cost_col = pick_column(saat, ["unit cost"], fallback_index=16)

    bin_code_col = pick_column(bins, ["item no"], fallback_index=0)

    bol_validation = validate_code_series(bol[bol_code_col], "BOL")
    saat_validation = validate_code_series(bins[bin_code_col], "SAAT")
    valid_bol_codes = set(bol_validation.loc[bol_validation["Status"].str.startswith("VALID"), "Clean Code"])
    valid_saat_codes = set(saat_validation.loc[saat_validation["Status"].str.startswith("VALID"), "Clean Code"])

    bol_by_code = bol.copy()
    bol_by_code["_code"] = bol_by_code[bol_code_col].map(clean_code)
    bol_by_code = bol_by_code[bol_by_code["_code"].isin(valid_bol_codes)].drop_duplicates("_code")

    saat_by_code = saat.copy()
    saat_by_code["_code"] = saat_by_code[saat_code_col].map(clean_code)
    saat_by_code = saat_by_code[saat_by_code["_code"].isin(valid_saat_codes)].drop_duplicates("_code")

    bol_item = first_lookup(bol_by_code, "_code", bol_item_col)
    bol_chemical = first_lookup(bol_by_code, "_code", bol_chemical_col)
    bol_price = first_lookup(bol_by_code, "_code", bol_price_col)
    bol_cas = first_lookup(bol_by_code, "_code", bol_cas_col)
    bol_appearance = first_lookup(bol_by_code, "_code", bol_appearance_col)
    saat_item = first_lookup(saat_by_code, "_code", saat_item_col)
    saat_cost = first_lookup(saat_by_code, "_code", saat_cost_col)

    chem_reference = (
        bol[[bol_chemical_col, bol_cas_col, bol_appearance_col]]
        .rename(columns={bol_chemical_col: "Chemical Name", bol_cas_col: "CAS Number", bol_appearance_col: "Appearance"})
        .dropna(subset=["Chemical Name"])
        .copy()
    )
    chem_reference["Chemical Name"] = chem_reference["Chemical Name"].map(lambda value: clean_text(value).upper())
    chem_reference = chem_reference[chem_reference["Chemical Name"] != ""].drop_duplicates("Chemical Name")

    manual_map: dict[str, str] = {}
    manual_cas_map: dict[str, str] = {}
    manual_appearance_map: dict[str, str] = {}
    if manual_chemicals is not None and not manual_chemicals.empty:
        manual_input = manual_chemicals.copy()
        if "Manual Chemical Name" not in manual_input.columns:
            manual_input["Manual Chemical Name"] = ""
        manual_input["_selected_chemical"] = manual_input["Chemical Name"].map(clean_text)
        manual_input["_manual_chemical"] = manual_input["Manual Chemical Name"].map(clean_text)
        manual_input["_final_chemical"] = manual_input["_manual_chemical"].where(
            manual_input["_manual_chemical"] != "",
            manual_input["_selected_chemical"],
        )
        manual_map = dict(zip(manual_input["Item Code"].map(clean_code), manual_input["_final_chemical"]))
        if "CAS Number" in manual_input.columns:
            cas_input = manual_input[manual_input["CAS Number"].map(clean_text) != ""]
            manual_cas_map = dict(zip(cas_input["Item Code"].map(clean_code), cas_input["CAS Number"].map(clean_text)))
        if "Appearance" in manual_input.columns:
            appearance_input = manual_input[
                ~manual_input["Appearance"].map(clean_text).isin(["", "0"])
            ]
            manual_appearance_map = dict(
                zip(appearance_input["Item Code"].map(clean_code), appearance_input["Appearance"].map(clean_text))
            )

    rows: list[dict[str, object]] = []
    for code in sorted(valid_bol_codes | valid_saat_codes):
        in_saat = code in valid_saat_codes
        in_bol = code in valid_bol_codes
        saat_mark = "*" if in_saat else ""
        bol_mark = "*" if in_bol else ""
        chemical_raw = clean_text(bol_chemical.get(code, "")) if in_bol else clean_text(manual_map.get(code, ""))
        chemical = chemical_raw.upper()

        if in_saat:
            item_name = clean_text(saat_item.get(code, "")).upper()
        else:
            item_name = clean_text(bol_item.get(code, "")).upper()

        if in_bol:
            price = to_number(bol_price.get(code, 0))
            cas_number = clean_text(bol_cas.get(code, ""))
            appearance = clean_text(bol_appearance.get(code, ""))
        else:
            price = to_number(saat_cost.get(code, 0)) / usd_to_idr if usd_to_idr else 0
            match = chem_reference[chem_reference["Chemical Name"] == chemical]
            cas_number = clean_text(match["CAS Number"].iloc[0]) if chemical and not match.empty else ""
            appearance = clean_text(match["Appearance"].iloc[0]) if chemical and not match.empty else "0"
            cas_number = manual_cas_map.get(code, "") or cas_number
            appearance = manual_appearance_map.get(code, "") or appearance

        rows.append(
            {
                "Item Code": code,
                "SAAT": saat_mark,
                "BOL": bol_mark,
                "Jumlah *": f"{saat_mark}{bol_mark}",
                "Item Name": item_name,
                "Chemical Name": chemical,
                "Price (USD)/KG": price,
                "CAS Number": cas_number,
                "Appearance": appearance,
            }
        )

    output = pd.DataFrame(rows)
    if not output.empty:
        output["_sort_chemical"] = output["Chemical Name"].replace("", pd.NA)
        output = (
            output.sort_values(["_sort_chemical", "Item Name", "Item Code"], na_position="last")
            .drop(columns="_sort_chemical")
            .reset_index(drop=True)
        )

    metrics = {
        "valid_saat": int(len(valid_saat_codes)),
        "valid_bol": int(len(valid_bol_codes)),
        "both": int(len(valid_saat_codes & valid_bol_codes)),
        "saat_only": int(len(valid_saat_codes - valid_bol_codes)),
        "bol_only": int(len(valid_bol_codes - valid_saat_codes)),
    }
    return output, bol_validation, saat_validation, metrics


def build_summary(validation: pd.DataFrame) -> pd.DataFrame:
    valid = validation[validation["Status"].str.startswith("VALID")]
    summary = (
        valid.groupby(["Source", "Segment"], as_index=False)
        .size()
        .pivot(index="Segment", columns="Source", values="size")
        .reindex(VALID_SEGMENTS)
        .fillna(0)
        .astype(int)
        .reset_index()
    )
    for col in ["SAAT", "BOL"]:
        if col not in summary.columns:
            summary[col] = 0
    return summary[["Segment", "SAAT", "BOL"]]


def render_section(title: str, subtitle: str) -> None:
    st.markdown(
        f"""
        <div class="section-kicker">{title}</div>
        <div class="mini-note">{subtitle}</div>
        """,
        unsafe_allow_html=True,
    )


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=OUTPUT_SHEET)
        worksheet = writer.sheets[OUTPUT_SHEET]
        widths = {
            "A": 16,
            "B": 8,
            "C": 8,
            "D": 10,
            "E": 52,
            "F": 42,
            "G": 16,
            "H": 16,
            "I": 14,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.style = "Headline 4"
    return buffer.getvalue()


def build_upload_template_bytes() -> bytes:
    """Create an upload template with the exact sheets and required columns."""
    bol_template = pd.DataFrame(
        [
            {
                "Chemical Name": "(example) Beta Pinene",
                "Commercial Price (USD) / Kg": 9,
                "Material Code [Segment 1+2+3+4]": "CF-BETA-PINENE-LE.ES-ALL",
                "CAS No.": "18172-67-3",
                "NAV Code": "GEN-CF-00817",
                "Physical State": "LIQUID",
            }
        ]
    )
    saat_template = pd.DataFrame(
        [
            {
                "No.": "GEN-CF-00901",
                "Description": "CF-(-)-BETA-PINENE >=97%, FCC, FG-M.ID-ID",
                "Unit Cost": 628250,
            }
        ]
    )
    bin_template = pd.DataFrame(
        [
            {
                "Item No.": "GEN-CF-00901",
                "Item Description": "CF-(-)-BETA-PINENE >=97%, FCC, FG-M.ID-ID",
                "Quantity": 1,
                "Unit of Measure Code": "KG",
                "Location Code": "PTSAAT-SA1",
                "Bin Code": "A-A-1-1",
                "Bin Description": "PTSAAT-SA1-A-A-1-1",
            }
        ]
    )
    guide = pd.DataFrame(
        [
            ["01 BOL DB", "NAV Code", "BOL code in GEN-XXX-00000 format"],
            ["01 BOL DB", "Chemical Name", "Reference list for the dropdown"],
            ["01 BOL DB", "Commercial Price (USD) / Kg", "BOL price in USD per kg"],
            ["01 BOL DB", "Material Code [Segment 1+2+3+4]", "Item name for BOL-only rows"],
            ["01 BOL DB", "CAS No.", "CAS number for the chemical name"],
            ["01 BOL DB", "Physical State", "Appearance example: LIQUID / SOLID"],
            ["01 SAAT DB", "No.", "SAAT item code"],
            ["01 SAAT DB", "Description", "Item name for SAAT rows"],
            ["01 SAAT DB", "Unit Cost", "IDR per kg converted using the exchange rate"],
            ["02 SAAT BIN", "Item No.", "SAAT item code from bin; keep the header on row 2"],
        ],
        columns=["Sheet", "Column", "Description"],
    )

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        guide.to_excel(writer, index=False, sheet_name="Instructions")
        bol_template.to_excel(writer, index=False, sheet_name="01 BOL DB")
        saat_template.to_excel(writer, index=False, sheet_name="01 SAAT DB")
        bin_template.to_excel(writer, index=False, sheet_name="02 SAAT BIN", startrow=1)

        workbook = writer.book
        bin_sheet = writer.sheets["02 SAAT BIN"]
        bin_sheet["A1"] = "Edit - Bin Contents - A-A-1-1"

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True, color="1F2933")
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            if worksheet.title == "02 SAAT BIN":
                worksheet.freeze_panes = "A3"
                for cell in worksheet[2]:
                    cell.fill = header_fill
                    cell.font = header_font

        widths_by_sheet = {
            "Instructions": {"A": 16, "B": 34, "C": 56},
            "01 BOL DB": {"A": 28, "B": 22, "C": 42, "D": 16, "E": 16, "F": 16},
            "01 SAAT DB": {"A": 16, "B": 48, "C": 14},
            "02 SAAT BIN": {"A": 16, "B": 48, "C": 12, "D": 20, "E": 18, "F": 14, "G": 24},
        }
        for sheet_name, widths in widths_by_sheet.items():
            worksheet = writer.sheets[sheet_name]
            for column, width in widths.items():
                worksheet.column_dimensions[column].width = width

    return buffer.getvalue()


def main() -> None:
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

    required_sheets = ["01 BOL DB", "01 SAAT DB", "02 SAAT BIN"]

    with st.sidebar:
        st.markdown("**Clean Database**")
        st.markdown(
            "<div class='mini-note'>Download a template, upload the workbook, and keep the matching step manual when needed.</div>",
            unsafe_allow_html=True,
        )
        st.download_button(
            "Download Upload Template",
            data=build_upload_template_bytes(),
            file_name="template-upload-bol-saat.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        uploaded = st.file_uploader("Upload workbook", type=["xlsx"])
        source = uploaded if uploaded is not None else DEFAULT_WORKBOOK
        usd_to_idr = st.number_input("USD to IDR exchange rate", min_value=1.0, value=16000.0, step=100.0)
        st.markdown("<div class='mini-note'>Required sheets</div>", unsafe_allow_html=True)
        available_sheets = set(read_sheet_names(source)) if uploaded is not None or DEFAULT_WORKBOOK.exists() else set()
        for sheet in required_sheets:
            left, right = st.columns([3, 1], gap="small")
            left.markdown(f"<div class='sheet-status-name'>{sheet}</div>", unsafe_allow_html=True)
            right.markdown(render_sheet_status_pill(sheet in available_sheets), unsafe_allow_html=True)

    if uploaded is None:
        source_key = f"default:{DEFAULT_WORKBOOK.stat().st_size if DEFAULT_WORKBOOK.exists() else 0}:{usd_to_idr}"
    else:
        source_key = f"upload:{uploaded.name}:{uploaded.size}:{usd_to_idr}"
    if st.session_state.get("source_key") != source_key:
        st.session_state["source_key"] = source_key
        st.session_state["final_generated"] = False
        st.session_state.pop("manual_saat_only", None)

    st.markdown(
        """
        <div class="hero">
          <div class="eyebrow">SAAT - <em>Make It Happen</em></div>
          <h1>BOL x SAAT database cleaner</h1>
          <p>Validate item codes, resolve chemicals, and export a calm, auditable workbook with a clean, Apple-like interface.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if uploaded is None and not DEFAULT_WORKBOOK.exists():
        st.warning("Upload an Excel file or place Bahan.xlsx in the app folder.")
        st.stop()

    try:
        bol, saat, bins = read_workbook(source)
        initial_output, bol_validation, saat_validation, metrics = build_output(bol, saat, bins, usd_to_idr)
    except Exception as exc:
        st.error(f"Failed to read workbook: {exc}")
        st.stop()

    only_saat = initial_output[(initial_output["SAAT"] == "*") & (initial_output["BOL"] == "")].copy()
    only_saat["Manual Chemical Name"] = ""
    only_saat["CAS Number"] = ""
    only_saat["Appearance"] = ""
    editable_cols = [
        "Item Code",
        "Item Name",
        "Chemical Name",
        "Manual Chemical Name",
        "Price (USD)/KG",
        "CAS Number",
        "Appearance",
    ]
    chemical_options = [""] + sorted(
        clean_text(value)
        for value in bol[pick_column(bol, ["chemical name"], fallback_index=1)].dropna().unique().tolist()
        if clean_text(value)
    )

    render_section("01 Overview", "Validation status and segment summary appear here before manual curation.")
    cols = st.columns(5)
    cols[0].metric("SAAT Valid", metrics["valid_saat"])
    cols[1].metric("BOL Valid", metrics["valid_bol"])
    cols[2].metric("In Both", metrics["both"])
    cols[3].metric("SAAT Only", metrics["saat_only"])
    cols[4].metric("BOL Only", metrics["bol_only"])

    summary = build_summary(pd.concat([saat_validation, bol_validation], ignore_index=True))
    st.dataframe(summary, use_container_width=True, hide_index=True)

    st.markdown("<div style='height:0.35rem'></div>", unsafe_allow_html=True)
    render_section("02 Manual curation", "For SAAT-only rows, choose a matching chemical name or fill it in manually when no reference exists.")

    with st.form("manual_saat_only_form"):
        edited_only_saat = st.data_editor(
            only_saat[editable_cols],
            use_container_width=True,
            hide_index=True,
            height=360,
            disabled=["Item Code", "Item Name", "Price (USD)/KG"],
            column_config={
                "Chemical Name": st.column_config.SelectboxColumn(
                    "Chemical Name",
                    options=chemical_options,
                    required=False,
                ),
                "Manual Chemical Name": st.column_config.TextColumn("Manual Chemical Name"),
                "CAS Number": st.column_config.TextColumn("CAS Number"),
                "Appearance": st.column_config.TextColumn("Appearance"),
            },
        )
        generate_clicked = st.form_submit_button("Generate Final Output", type="primary")

    if generate_clicked:
        st.session_state["manual_saat_only"] = edited_only_saat
        st.session_state["final_generated"] = True

    tab_validation, tab_output = st.tabs(["Validation Details", "Final Output"])
    with tab_validation:
        validation = pd.concat([saat_validation, bol_validation], ignore_index=True)
        status_filter = st.multiselect(
            "Filter status",
            sorted(validation["Status"].unique().tolist()),
            default=sorted(validation["Status"].unique().tolist()),
        )
        st.dataframe(
            validation[validation["Status"].isin(status_filter)],
            use_container_width=True,
            hide_index=True,
            height=520,
        )

    with tab_output:
        if not st.session_state.get("final_generated"):
            st.markdown(
                "<div class='soft-panel'><div class='mini-note'>Finish the manual matching for SAAT-only rows, then click <strong>Generate Final Output</strong>.</div></div>",
                unsafe_allow_html=True,
            )
            st.stop()

        final_output, _, _, metrics = build_output(
            bol,
            saat,
            bins,
            usd_to_idr,
            st.session_state.get("manual_saat_only", edited_only_saat),
        )
        st.caption(f"{len(final_output):,} rows".replace(",", "."))
        st.dataframe(final_output, use_container_width=True, hide_index=True, height=560)

        st.markdown("<div style='height:0.25rem'></div>", unsafe_allow_html=True)
        st.download_button(
            "Download Excel",
            data=to_excel_bytes(final_output),
            file_name="BOL-SAAT List.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()
