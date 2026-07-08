from copy import copy
import csv
import io
import json
from datetime import date as date_class, datetime
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
from typing import Any, Dict, Iterable, List, Optional

import openpyxl
import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path(__file__).resolve().parent
LOGO_PATH = PROJECT_ROOT / "assets" / "logo.png"
LOGO_MERGED_RANGE = "A1:B3"
EMPTY_SECTION_ROW_COUNT = 3


MAIN_PHASES = [
    "Casing Rajangan",
    "Casing Krosok",
    "Top Flavor",
]

DEFAULT_PREMIX_PHASES = [
    "Casing Pre-Mix",
    "Flavor Pre-Mix 1",
    "Flavor Pre-Mix 2",
]

KNOWN_PHASES = MAIN_PHASES + DEFAULT_PREMIX_PHASES

DYNAMIC_FORMULA_PHASES = KNOWN_PHASES.copy()

BASE_SECTION_ROW_RANGES = {
    "Casing Rajangan": list(range(26, 31)),
    "Casing Krosok": list(range(39, 44)),
    "Top Flavor": list(range(52, 54)),
    "Casing Pre-Mix": list(range(77, 87)),
    "Flavor Pre-Mix 1": list(range(93, 128)),
    "Flavor Pre-Mix 2": list(range(134, 143)),
}

SECTION_ROW_RANGES = {phase: rows.copy() for phase, rows in BASE_SECTION_ROW_RANGES.items()}

BASE_PHASE_METADATA_POSITIONS = {
    "Casing Rajangan": {"nav_code": (21, 5), "description": (22, 5), "blend_ratio": (23, 5), "application": (24, 5)},
    "Casing Krosok": {"nav_code": (34, 5), "description": (35, 5), "blend_ratio": (36, 5), "application": (37, 5)},
    "Top Flavor": {"nav_code": (47, 5), "description": (48, 5), "blend_ratio": (49, 5), "application": (50, 5)},
    "Casing Pre-Mix": {"nav_code": (74, 5), "description": (75, 5), "blend_ratio": None, "application": None},
    "Flavor Pre-Mix 1": {"nav_code": (90, 5), "description": (91, 5), "blend_ratio": None, "application": None},
    "Flavor Pre-Mix 2": {"nav_code": (131, 5), "description": (132, 5), "blend_ratio": None, "application": None},
}

PHASE_METADATA_POSITIONS = {
    phase: positions.copy() for phase, positions in BASE_PHASE_METADATA_POSITIONS.items()
}

FORMULA_METADATA_POSITIONS = {
    "product_name": (5, 3),
    "prepared_by": (7, 3),
    "effective_date": (3, 17),
    "standard_control": (10, 3),
    "flavor_standard_reference": (10, 10),
    "tobacco_blend_code": (12, 3),
    "sensory_parameter": (12, 10),
    "formulation_code": (13, 3),
    "impact": (13, 10),
    "single_capsule": (14, 3),
    "flavor_aroma": (14, 10),
    "double_capsule_tobacco_end": (15, 3),
    "irritation": (15, 10),
    "double_capsule_mouth_end": (16, 3),
    "cooling": (16, 10),
    "product_weight_mg_stick": (18, 3),
    "clove_weight_mg_stick": (18, 6),
    "stick_per_mc": (18, 10),
}

INPUT_METADATA_POSITIONS = {
    "product_name": (2, 2),
    "formula_code": (3, 2),
    "product_weight_mg_stick": (4, 2),
    "clove_weight_mg_stick": (5, 2),
    "stick_per_mc": (6, 2),
    "prepared_by": (7, 2),
    "date": (8, 2),
    "standard_control": (3, 11),
    "flavor_standard_reference": (3, 14),
    "tobacco_blend_code": (5, 11),
    "sensory_parameter": (5, 14),
    "formulation_code": (6, 11),
    "impact": (6, 14),
    "single_capsule": (7, 11),
    "flavor_aroma": (7, 14),
    "double_capsule_tobacco_end": (8, 11),
    "irritation": (8, 14),
    "double_capsule_mouth_end": (9, 11),
    "cooling": (9, 14),
}

PHASE_METADATA_START_ROW = 4
PHASE_METADATA_ROW_COUNT = 14
MATERIAL_HEADER_ROW = 19
MATERIAL_START_ROW = 20
MATERIAL_COLUMNS = {
    "phase": 1,
    "item_code": 2,
    "item_name": 3,
    "dosage_input_mode": 4,
    "dosage_mg_stick": 5,
    "ratio_percent": 6,
    "addition_sequence": 7,
    "temperature": 8,
    "agitation_rate": 9,
    "mixing_duration": 10,
    "work_instruction_override": 11,
    "process_role": 12,
    "notes": 13,
}

FORMULA_MATERIAL_COLUMNS = {
    "no": 1,
    "item_code": 2,
    "item_name": 3,
    "physical_form": 5,
    "cas_number": 6,
    "ratio": 7,
    "dosage_mg_stick": 8,
    "material_price": 9,
    "formulation_price": 10,
    "dosage_kg_mc": 11,
    "density": 12,
    "addition_sequence": 13,
    "temperature": 14,
    "agitation_rate": 15,
    "mixing_duration": 16,
    "work_instruction": 17,
}

REQUIRED_FORMULATION_FIELDS = [
    "product_name",
    "formula_code",
    "prepared_by",
    "date",
    "product_weight_mg_stick",
    "clove_weight_mg_stick",
    "stick_per_mc",
]


@dataclass
class PhaseMetadata:
    phase: str
    nav_code: Optional[str] = None
    description: Optional[str] = None
    blend_ratio: Optional[float] = None
    application: Optional[float] = None


@dataclass
class MaterialInput:
    phase: str
    item_code: Optional[str]
    item_name: Optional[str]
    dosage_mg_stick: Optional[float]
    dosage_input_mode: str = "mg/stick"
    ratio_percent: Optional[float] = None
    application_percent: Optional[float] = None
    addition_sequence: Optional[int] = None
    temperature: Optional[str] = None
    agitation_rate: Optional[str] = None
    mixing_duration: Optional[str] = None
    work_instruction_override: Optional[str] = None
    process_role: Optional[str] = None
    notes: Optional[str] = None
    physical_form: Optional[str] = None
    cas_number: Optional[str] = None
    material_price: Optional[float] = None
    ratio: Optional[float] = None
    formulation_price: Optional[float] = None
    dosage_kg_mc: Optional[float] = None
    work_instruction: Optional[str] = None


@dataclass
class FormulationInput:
    product_name: str
    formula_code: str
    product_weight_mg_stick: float
    clove_weight_mg_stick: float
    stick_per_mc: int
    prepared_by: str
    date: Any
    standard_control: str
    flavor_standard_reference: str
    tobacco_blend_code: str
    sensory_parameter: str
    formulation_code: str
    impact: str
    single_capsule: str
    flavor_aroma: str
    double_capsule_tobacco_end: str
    irritation: str
    double_capsule_mouth_end: str
    cooling: str
    phase_metadata: Dict[str, PhaseMetadata] = field(default_factory=dict)
    materials: List[MaterialInput] = field(default_factory=list)

    @property
    def effective_date(self) -> Any:
        return self.date


def normalize_phase(phase: str) -> str:
    if phase is None:
        raise ValueError("Phase name is required")
    phase_text = str(phase).strip()
    for known in KNOWN_PHASES:
        if phase_text.lower() == known.lower():
            return known
    lower_phase = phase_text.lower()
    for prefix in ("flavor pre-mix", "casing pre-mix"):
        if not lower_phase.startswith(prefix):
            continue
        suffix = phase_text[len(prefix):].strip()
        if not suffix:
            return "Casing Pre-Mix" if prefix == "casing pre-mix" else "Flavor Pre-Mix 1"
        try:
            number = int(suffix)
        except ValueError:
            break
        if number >= 1:
            if prefix == "casing pre-mix":
                return "Casing Pre-Mix" if number == 1 else f"Casing Pre-Mix {number}"
            return f"Flavor Pre-Mix {number}"
    return phase_text


def is_premix_phase(phase: str) -> bool:
    normalized = normalize_phase(phase)
    return normalized.startswith("Casing Pre-Mix") or normalized.startswith("Flavor Pre-Mix")


def phase_sort_key(phase: str) -> tuple:
    normalized = normalize_phase(phase)
    if normalized in MAIN_PHASES:
        return (0, MAIN_PHASES.index(normalized))
    if normalized.startswith("Casing Pre-Mix"):
        suffix = normalized.replace("Casing Pre-Mix", "").strip()
        number = int(suffix) if suffix.isdigit() else 1
        return (1, number)
    if normalized.startswith("Flavor Pre-Mix"):
        suffix = normalized.replace("Flavor Pre-Mix", "").strip()
        number = int(suffix) if suffix.isdigit() else 1
        return (2, number)
    return (9, normalized)


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def is_missing_number(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def normalize_input_mode(value: Any) -> str:
    text = str(value or "mg/stick").strip().lower()
    if text in ("%", "percent", "percentage", "ratio/application %", "ratio %"):
        return "%"
    return "mg/stick"


def percent_to_fraction(value: Any) -> float:
    number = float(value)
    if abs(number) > 1:
        return number / 100
    return number


def round_dosage(value: float) -> float:
    return round(float(value), 5)


def effective_material_application_percent(formulation: FormulationInput, material: MaterialInput) -> Optional[float]:
    if material.application_percent not in (None, ""):
        return material.application_percent
    phase_metadata = formulation.phase_metadata.get(normalize_phase(material.phase))
    if phase_metadata and phase_metadata.application not in (None, ""):
        return phase_metadata.application
    return None


def reset_runtime_layout() -> None:
    SECTION_ROW_RANGES.clear()
    SECTION_ROW_RANGES.update({phase: rows.copy() for phase, rows in BASE_SECTION_ROW_RANGES.items()})
    PHASE_METADATA_POSITIONS.clear()
    PHASE_METADATA_POSITIONS.update({
        phase: positions.copy() for phase, positions in BASE_PHASE_METADATA_POSITIONS.items()
    })
    DYNAMIC_FORMULA_PHASES.clear()
    DYNAMIC_FORMULA_PHASES.extend(KNOWN_PHASES)


def shift_runtime_layout(insert_at: int, row_count: int, skip_phase: Optional[str] = None) -> None:
    for phase, rows in list(SECTION_ROW_RANGES.items()):
        if phase == skip_phase:
            continue
        SECTION_ROW_RANGES[phase] = [row + row_count if row >= insert_at else row for row in rows]

    for phase, positions in PHASE_METADATA_POSITIONS.items():
        shifted: Dict[str, Optional[tuple]] = {}
        for field_name, position in positions.items():
            if position is None:
                shifted[field_name] = None
                continue
            row, column = position
            shifted[field_name] = (row + row_count, column) if row >= insert_at else position
        PHASE_METADATA_POSITIONS[phase] = shifted


def desired_section_row_count(rows: List["MaterialInput"]) -> int:
    if rows:
        return len(rows)
    return EMPTY_SECTION_ROW_COUNT


def load_material_db(template_path: Path) -> pd.DataFrame:
    df = pd.read_excel(template_path, sheet_name="BOL-SAAT List", engine="openpyxl")
    columns = {col: col for col in df.columns}
    normalized = {
        "item_code": columns.get("Item Code", "Item Code"),
        "item_name": columns.get("Item Name", "Item Name"),
        "price": columns.get("Price (USD)/KG", "Price (USD)/KG"),
        "cas_number": columns.get("CAS Number", "CAS Number"),
        "appearance": columns.get("Appearance", "Appearance"),
    }
    df = df.rename(columns={
        normalized["item_code"]: "item_code",
        normalized["item_name"]: "item_name",
        normalized["price"]: "price",
        normalized["cas_number"]: "cas_number",
        normalized["appearance"]: "appearance",
    })
    df = df.loc[:, ["item_code", "item_name", "price", "cas_number", "appearance"]]
    df["item_code"] = df["item_code"].astype(str).str.strip()
    df["item_name"] = df["item_name"].astype(str).str.strip()
    return df


def lookup_material_record(df: pd.DataFrame, item_code: Optional[str], item_name: Optional[str]) -> Optional[Dict[str, Any]]:
    if item_code:
        code = str(item_code).strip()
        row = df.loc[df["item_code"].str.lower() == code.lower()]
        if not row.empty:
            row = row.iloc[0]
            return row.to_dict()

    if item_name:
        name = str(item_name).strip().lower()
        row = df.loc[df["item_name"].str.lower() == name]
        if not row.empty:
            row = row.iloc[0]
            return row.to_dict()

    return None


def find_material_record(df: pd.DataFrame, item_code: Optional[str], item_name: Optional[str]) -> Dict[str, Any]:
    record = lookup_material_record(df, item_code, item_name)
    if record is not None:
        return record
    return {"item_code": item_code, "item_name": item_name, "price": None, "cas_number": None, "appearance": None}


def validate_formulation_input(formulation: FormulationInput, material_db: pd.DataFrame) -> List[str]:
    errors: List[str] = []

    for field_name in REQUIRED_FORMULATION_FIELDS:
        value = getattr(formulation, field_name)
        if is_blank(value):
            errors.append(f"Required field '{field_name}' wajib diisi.")

    try:
        if float(formulation.stick_per_mc) <= 0:
            errors.append("Field 'stick_per_mc' harus lebih dari 0.")
    except (TypeError, ValueError):
        errors.append("Field 'stick_per_mc' harus berupa angka.")

    if not formulation.materials:
        errors.append("Minimal satu material wajib diisi.")

    for index, material in enumerate(formulation.materials, start=1):
        row_label = f"Material row {index}"
        phase = normalize_phase(material.phase) if not is_blank(material.phase) else ""
        if not phase:
            errors.append(f"{row_label}: phase wajib diisi.")
        elif phase not in MAIN_PHASES and not is_premix_phase(phase):
            errors.append(
                f"{row_label}: phase '{phase}' tidak didukung. "
                "Gunakan Casing Rajangan, Casing Krosok, Top Flavor, atau pola Casing/Flavor Pre-Mix."
            )

        input_mode = normalize_input_mode(material.dosage_input_mode)
        if input_mode == "%":
            try:
                ratio = percent_to_fraction(material.ratio_percent)
                application = percent_to_fraction(effective_material_application_percent(formulation, material))
                if ratio <= 0 or application <= 0:
                    errors.append(f"{row_label}: ratio_percent dan phase application harus lebih dari 0 untuk mode %.")
            except (TypeError, ValueError):
                errors.append(f"{row_label}: ratio_percent wajib diisi dan Application % di Phase Metadata wajib angka untuk mode %.")
        else:
            try:
                dosage = float(material.dosage_mg_stick)
                if dosage <= 0:
                    errors.append(f"{row_label}: dosage_mg_stick harus lebih dari 0.")
            except (TypeError, ValueError):
                errors.append(f"{row_label}: dosage_mg_stick harus berupa angka untuk mode mg/stick.")

        if is_blank(material.item_code) and is_blank(material.item_name):
            errors.append(f"{row_label}: item_code atau item_name wajib diisi.")
            continue

        record = lookup_material_record(material_db, material.item_code, material.item_name)
        if record is None:
            lookup_value = material.item_code or material.item_name
            errors.append(f"{row_label}: material lookup tidak ditemukan untuk '{lookup_value}'.")
            continue

        price = material.material_price if material.material_price is not None else record.get("price")
        if is_missing_number(price):
            lookup_value = material.item_code or material.item_name
            errors.append(f"{row_label}: price material kosong untuk '{lookup_value}'.")

    phases_with_materials = {normalize_phase(material.phase) for material in formulation.materials if not is_blank(material.phase)}
    for phase in MAIN_PHASES:
        if phase not in phases_with_materials:
            errors.append(f"Phase utama '{phase}' wajib memiliki minimal satu material.")

    return errors


def build_work_instruction(material: MaterialInput, phase_first: bool) -> str:
    if material.work_instruction_override:
        return material.work_instruction_override
    physical = material.physical_form
    if physical is None:
        physical = ""
    elif not isinstance(physical, str):
        physical = str(physical)
    physical = physical.strip().lower()
    if phase_first:
        if "solid" in physical:
            return "Magnetic stirrer. Add solids and mix until fully dissolved."
        return "Magnetic stirrer. Start mixing base materials till uniform."
    if material.process_role:
        return f"{material.process_role}. Continue mixing until homogeneous."
    if material.addition_sequence is not None:
        return f"Add material in sequence {material.addition_sequence}. Mix until uniform."
    return "Magnetic stirrer. Mix until homogeneous."


def compute_materials(formulation: FormulationInput, material_db: pd.DataFrame) -> None:
    phase_groups: Dict[str, List[MaterialInput]] = {}
    for material in formulation.materials:
        material.phase = normalize_phase(material.phase)
        material.dosage_input_mode = normalize_input_mode(material.dosage_input_mode)
        if material.dosage_input_mode == "%":
            tobacco_weight = float(formulation.product_weight_mg_stick) - float(formulation.clove_weight_mg_stick)
            ratio = percent_to_fraction(material.ratio_percent)
            application = percent_to_fraction(effective_material_application_percent(formulation, material))
            material.dosage_mg_stick = round_dosage(ratio * application * tobacco_weight)
        else:
            material.dosage_mg_stick = round_dosage(float(material.dosage_mg_stick or 0))
        phase_groups.setdefault(material.phase, []).append(material)

    for phase, materials in phase_groups.items():
        total_dosage = sum((m.dosage_mg_stick or 0) for m in materials)
        materials_sorted = sorted(materials, key=lambda m: (m.addition_sequence if m.addition_sequence is not None else 999, m.item_code or ""))
        for index, material in enumerate(materials_sorted, start=1):
            record = find_material_record(material_db, material.item_code, material.item_name)
            material.physical_form = material.physical_form or record.get("appearance")
            material.cas_number = material.cas_number or record.get("cas_number")
            material.material_price = material.material_price if material.material_price is not None else record.get("price")
            material.item_code = material.item_code or record.get("item_code")
            material.item_name = material.item_name or record.get("item_name")
            material.addition_sequence = material.addition_sequence or index
            material.ratio = 0.0 if total_dosage <= 0 else (material.dosage_mg_stick or 0) / total_dosage
            material.formulation_price = (material.ratio or 0.0) * (material.material_price or 0.0)
            material.dosage_kg_mc = (material.dosage_mg_stick or 0.0) * formulation.stick_per_mc / 1_000_000
            material.work_instruction = build_work_instruction(material, index == 1)

    formulation.materials = sorted(
        formulation.materials,
        key=lambda material: (
            phase_sort_key(material.phase),
            material.addition_sequence if material.addition_sequence is not None else 999999,
            str(material.item_code or material.item_name or ""),
        ),
    )


def writable_cell(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, column: int):
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return None
    return cell


def set_cell_value(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, column: int, value: Any) -> None:
    cell = writable_cell(ws, row, column)
    if cell is not None:
        cell.value = value


def clear_cells(ws: openpyxl.worksheet.worksheet.Worksheet, rows: Iterable[int], columns: Iterable[int]) -> None:
    for row in rows:
        for column in columns:
            set_cell_value(ws, row, column, None)


def copy_cell_format(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def copy_material_row_style(ws: openpyxl.worksheet.worksheet.Worksheet, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=column)
        target = ws.cell(row=target_row, column=column)
        if not isinstance(target, MergedCell):
            copy_cell_format(source, target)

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == source_row and merged_range.max_row == source_row:
            ws.merge_cells(
                start_row=target_row,
                start_column=merged_range.min_col,
                end_row=target_row,
                end_column=merged_range.max_col,
            )


def copy_row_content_and_style(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    source_row: int,
    target_row: int,
) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].hidden = False
    for column in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=column)
        target = ws.cell(row=target_row, column=column)
        if isinstance(source, MergedCell) or isinstance(target, MergedCell):
            continue
        target.value = source.value
        copy_cell_format(source, target)


def copy_merged_ranges_for_block(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    source_start: int,
    source_end: int,
    target_start: int,
) -> None:
    row_offset = target_start - source_start
    existing_ranges = {str(merged_range) for merged_range in ws.merged_cells.ranges}
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row < source_start or merged_range.max_row > source_end:
            continue
        target_range = CellRange(
            min_col=merged_range.min_col,
            min_row=merged_range.min_row + row_offset,
            max_col=merged_range.max_col,
            max_row=merged_range.max_row + row_offset,
        )
        if str(target_range) not in existing_ranges:
            ws.merge_cells(str(target_range))
            existing_ranges.add(str(target_range))


def premix_source_phase(phase: str) -> str:
    return "Casing Pre-Mix" if normalize_phase(phase).startswith("Casing Pre-Mix") else "Flavor Pre-Mix 2"


def premix_insert_row(phase: str) -> int:
    normalized = normalize_phase(phase)
    if normalized.startswith("Casing Pre-Mix"):
        flavor_positions = [
            positions["nav_code"][0] - 1
            for phase_name, positions in PHASE_METADATA_POSITIONS.items()
            if phase_name.startswith("Flavor Pre-Mix") and positions.get("nav_code")
        ]
        if flavor_positions:
            return min(flavor_positions)
    premix_phases = [
        phase_name
        for phase_name in SECTION_ROW_RANGES
        if is_premix_phase(phase_name)
    ]
    return max(SECTION_ROW_RANGES[phase_name][-1] + 2 for phase_name in premix_phases)


def ensure_dynamic_premix_layouts(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    phase_rows: Dict[str, List[MaterialInput]],
) -> None:
    requested_phases = sorted(
        {normalize_phase(phase) for phase in phase_rows if is_premix_phase(phase)},
        key=phase_sort_key,
    )
    for phase in requested_phases:
        if phase in SECTION_ROW_RANGES:
            continue

        source_phase = premix_source_phase(phase)
        source_rows = SECTION_ROW_RANGES[source_phase]
        source_start = PHASE_METADATA_POSITIONS[source_phase]["nav_code"][0] - 1
        source_end = source_rows[-1] + 1
        block_height = source_end - source_start + 1
        insert_at = premix_insert_row(phase)

        ws.insert_rows(insert_at, amount=block_height)
        shift_runtime_layout(insert_at, block_height)

        shifted_source_start = source_start + block_height if source_start >= insert_at else source_start
        shifted_source_end = source_end + block_height if source_end >= insert_at else source_end
        for offset in range(block_height):
            copy_row_content_and_style(ws, shifted_source_start + offset, insert_at + offset)
        copy_merged_ranges_for_block(ws, shifted_source_start, shifted_source_end, insert_at)

        new_rows = [
            insert_at + (row - shifted_source_start)
            for row in range(shifted_source_start, shifted_source_end + 1)
            if shifted_source_start + 4 <= row <= shifted_source_end - 1
        ]
        SECTION_ROW_RANGES[phase] = new_rows
        PHASE_METADATA_POSITIONS[phase] = {
            "nav_code": (insert_at + 1, 5),
            "description": (insert_at + 2, 5),
            "blend_ratio": None,
            "application": None,
        }
        DYNAMIC_FORMULA_PHASES.append(phase)

        label_prefix = "Casing Pre-Mix Formulation" if phase.startswith("Casing Pre-Mix") else "Flavor Pre-Mix Formulation"
        set_cell_value(ws, insert_at, 1, label_prefix)
        set_cell_value(ws, insert_at + 1, 1, f"{label_prefix} NAV Item Code")
        set_cell_value(ws, insert_at + 2, 1, f"{label_prefix} NAV Item Description")


def unmerge_blocking_material_ranges(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    material_rows = {
        row
        for rows in SECTION_ROW_RANGES.values()
        for row in rows
    }
    for merged_range in list(ws.merged_cells.ranges):
        overlaps_material_rows = any(merged_range.min_row <= row <= merged_range.max_row for row in material_rows)
        overlaps_material_columns = merged_range.min_col <= 17 and merged_range.max_col >= 1
        if not overlaps_material_rows or not overlaps_material_columns:
            continue
        if merged_range.min_row == merged_range.max_row:
            if merged_range.min_col == 3 and merged_range.max_col == 4:
                continue

        source = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        ws.unmerge_cells(str(merged_range))
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                target = ws.cell(row=row, column=column)
                copy_cell_format(source, target)


def unmerge_ranges_overlapping_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    first_row: int,
    last_row: int,
) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row < first_row or merged_range.min_row > last_row:
            continue
        source = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        ws.unmerge_cells(str(merged_range))
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                target = ws.cell(row=row, column=column)
                copy_cell_format(source, target)


def prepare_dynamic_formula_sections(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    phase_rows: Dict[str, List[MaterialInput]],
) -> None:
    for phase in DYNAMIC_FORMULA_PHASES:
        rows = phase_rows.get(phase, [])
        base_row_indices = SECTION_ROW_RANGES[phase]
        target_count = desired_section_row_count(rows)
        row_delta = target_count - len(base_row_indices)

        if row_delta > 0:
            insert_at = max(base_row_indices) + 1
            source_row = max(base_row_indices)
            ws.insert_rows(insert_at, amount=row_delta)
            for offset in range(row_delta):
                copy_material_row_style(ws, source_row, insert_at + offset)
            SECTION_ROW_RANGES[phase] = list(range(base_row_indices[0], base_row_indices[0] + target_count))
            shift_runtime_layout(insert_at, row_delta, skip_phase=phase)
            continue

        visible_rows = base_row_indices[:target_count]
        total_row = visible_rows[-1] + 1
        hidden_rows = [
            row
            for row in base_row_indices[target_count:]
            if row != total_row
        ]
        old_total_row = base_row_indices[-1] + 1
        if old_total_row != total_row:
            hidden_rows.append(old_total_row)
        SECTION_ROW_RANGES[phase] = visible_rows
        for row in hidden_rows:
            clear_cells(ws, [row], range(1, 18))
            ws.row_dimensions[row].hidden = True

    unmerge_blocking_material_ranges(ws)


def build_formulation_code_formula() -> str:
    description_refs = []
    for phase in ("Casing Rajangan", "Casing Krosok", "Top Flavor"):
        description_position = PHASE_METADATA_POSITIONS[phase]["description"]
        description_refs.append(f"E{description_position[0]}")
    return f'=CONCATENATE({description_refs[0]},"/",{description_refs[1]},"/",{description_refs[2]})'


def format_approval_date(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d %B %Y")
    try:
        parsed = pd.to_datetime(value)
        if not pd.isna(parsed):
            return parsed.strftime("%d %B %Y")
    except (TypeError, ValueError):
        pass
    return str(value)


def apply_document_header_layout(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    formulation: FormulationInput,
) -> None:
    header_values = {
        "P1": "Document No",
        "Q1": "BOL-ID-BOLL-SPEC-003",
        "P2": "Revision No",
        "Q2": "00",
        "P3": "Effective Date",
        "Q3": formulation.effective_date,
        "A4": "Factory Lab to Fill Up",
        "A5": "Full FG Description",
        "A6": "Mixing Factory",
        "C6": "PT SAAT",
        "A7": "Lab Personnel Involved",
        "A9": "Flavor Development Reference",
        "A10": "Standard Control",
        "H10": "Flavor Standard Reference",
        "A11": "Product Specification",
        "A12": "Tobacco Blend Code",
        "H12": "Sensory Parameter",
        "A13": "Formulation Code",
        "H13": "Impact",
        "A14": "Single Capsule",
        "H14": "Flavor Aroma",
        "A15": "Double Capsule (Tobacco End)",
        "H15": "Irritation",
        "A16": "Double Capsule (Mouth End)",
        "H16": "Cooling",
    }
    for cell_ref, value in header_values.items():
        ws[cell_ref] = value

    ws["Q1"].font = Font(bold=True, color="FF0000")
    ws["Q2"].number_format = "@"
    for cell_ref in ("P1", "P2", "P3", "A4", "A5", "A6", "A7", "A9", "A10", "H10", "A11", "A12", "H12", "A13", "H13", "A14", "H14", "A15", "H15", "A16", "H16"):
        ws[cell_ref].font = copy(ws[cell_ref].font)
        ws[cell_ref].font = Font(bold=True)


def apply_approval_block(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    formulation: FormulationInput,
) -> None:
    approval_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Prepared By":
            approval_row = row
            break
    if approval_row is None:
        return

    approval_date = format_approval_date(formulation.effective_date)
    blocks = [
        (1, "Prepared By", "Position : Flavourist", f"Name : {formulation.prepared_by}", f"Date : {approval_date}", "Signature : "),
        (6, "Reviewed By", "Position : Senior Manager - Flavourist", "Name : Mochamad Setyawan", f"Date : {approval_date}", "Signature:"),
        (13, "Approved By", "Position : Head, Flavour PDI", "Name : Andrew Yip", f"Date : {approval_date}", "Signature:"),
    ]
    for column, title, position, name, date, signature in blocks:
        values = [title, position, name, date, signature]
        for offset, value in enumerate(values):
            cell = ws.cell(row=approval_row + offset, column=column)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if offset == 0:
                cell.font = Font(bold=True)


def regenerate_total_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for phase, row_indices in SECTION_ROW_RANGES.items():
        first_row = row_indices[0]
        last_row = row_indices[-1]
        total_row = last_row + 1
        ws.row_dimensions[total_row].hidden = False
        clear_cells(ws, [total_row], range(1, 18))
        set_cell_value(ws, total_row, 1, "Total")
        set_cell_value(ws, total_row, 7, f"=SUM(G{first_row}:G{last_row})")
        set_cell_value(ws, total_row, 8, f"=SUM(H{first_row}:H{last_row})")
        set_cell_value(ws, total_row, 10, f"=SUM(J{first_row}:J{last_row})")
        set_cell_value(ws, total_row, 11, f"=SUM(K{first_row}:K{last_row})")


def apply_formula_number_formats(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    currency_format = '"$"#,##0.00'
    percent_format = "0.000%"
    dosage_format = "0.00000"
    white_fill = PatternFill(fill_type=None)
    total_fill = PatternFill("solid", fgColor="D9D9D9")

    for row_indices in SECTION_ROW_RANGES.values():
        for row in row_indices:
            for column in range(1, 18):
                ws.cell(row=row, column=column).fill = white_fill
            ws.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["ratio"]).number_format = percent_format
            ws.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"]).number_format = dosage_format
            ws.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["material_price"]).number_format = currency_format
            ws.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["formulation_price"]).number_format = currency_format
            ws.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["dosage_kg_mc"]).number_format = "0.00000"

        total_row = row_indices[-1] + 1
        for column in range(1, 18):
            ws.cell(row=total_row, column=column).fill = total_fill
        ws.cell(row=total_row, column=7).number_format = percent_format
        ws.cell(row=total_row, column=8).number_format = dosage_format
        ws.cell(row=total_row, column=10).number_format = currency_format
        ws.cell(row=total_row, column=11).number_format = "0.00000"


def first_non_blank(values: Iterable[Any]) -> Any:
    for value in values:
        if not is_blank(value):
            return value
    return None


def merge_process_instruction_columns(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    phase_rows: Dict[str, List[MaterialInput]],
) -> None:
    process_columns = [
        FORMULA_MATERIAL_COLUMNS["addition_sequence"],
        FORMULA_MATERIAL_COLUMNS["temperature"],
        FORMULA_MATERIAL_COLUMNS["agitation_rate"],
        FORMULA_MATERIAL_COLUMNS["mixing_duration"],
        FORMULA_MATERIAL_COLUMNS["work_instruction"],
    ]
    value_getters = {
        FORMULA_MATERIAL_COLUMNS["addition_sequence"]: lambda rows: rows[0].addition_sequence,
        FORMULA_MATERIAL_COLUMNS["temperature"]: lambda rows: first_non_blank(row.temperature for row in rows),
        FORMULA_MATERIAL_COLUMNS["agitation_rate"]: lambda rows: first_non_blank(row.agitation_rate for row in rows),
        FORMULA_MATERIAL_COLUMNS["mixing_duration"]: lambda rows: first_non_blank(row.mixing_duration for row in rows),
        FORMULA_MATERIAL_COLUMNS["work_instruction"]: lambda rows: (
            first_non_blank(row.work_instruction_override for row in rows)
            or first_non_blank(row.work_instruction for row in rows)
        ),
    }

    for phase, materials in phase_rows.items():
        row_indices = SECTION_ROW_RANGES.get(phase, [])
        if not row_indices or not materials:
            continue

        group_start = 0
        while group_start < len(materials):
            sequence = materials[group_start].addition_sequence
            group_end = group_start + 1
            while group_end < len(materials) and materials[group_end].addition_sequence == sequence:
                group_end += 1

            start_row = row_indices[group_start]
            end_row = row_indices[group_end - 1]
            grouped_materials = materials[group_start:group_end]

            for column in process_columns:
                cell = ws.cell(row=start_row, column=column)
                cell.value = value_getters[column](grouped_materials)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                if column == FORMULA_MATERIAL_COLUMNS["work_instruction"]:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )
                if end_row > start_row:
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=column,
                        end_row=end_row,
                        end_column=column,
                    )

            group_start = group_end


def update_top_formula_summary(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_row = None
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=2).value
        if isinstance(value, str) and "Casing & Top Flavor" in value:
            header_row = row
            break
    if header_row is None:
        return

    clear_cells(ws, range(header_row, header_row + 5), range(1, 9))

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    total_fill = PatternFill("solid", fgColor="D9D9D9")
    white_fill = PatternFill(fill_type=None)
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    headers = {
        1: "No",
        2: "Casing & Top Flavor\nNAV Item Code",
        3: "Application Amount\n(mg/stick)",
        4: "Blend Ratio",
        7: "Price\n(USD/KG)",
        8: "1,000 Sticks Price\n(USD)",
    }
    for column, value in headers.items():
        cell = ws.cell(row=header_row, column=column)
        cell.value = value
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.number_format = "General"

    data_rows = []
    for index, phase in enumerate(("Casing Rajangan", "Casing Krosok", "Top Flavor"), start=1):
        row = header_row + index
        data_rows.append(row)
        positions = PHASE_METADATA_POSITIONS[phase]
        description = positions["description"]
        blend_ratio = positions["blend_ratio"]
        total_row = SECTION_ROW_RANGES[phase][-1] + 1

        set_cell_value(ws, row, 1, index)
        set_cell_value(ws, row, 2, f"=E{description[0]}")
        set_cell_value(ws, row, 3, f"=H{total_row}")
        set_cell_value(ws, row, 4, f"=E{blend_ratio[0]}")
        set_cell_value(ws, row, 7, f"=J{total_row}")
        set_cell_value(ws, row, 8, f"=G{row}*(C{row}/1000000)*1000")

        for column in (1, 2, 3, 4, 7, 8):
            cell = ws.cell(row=row, column=column)
            cell.fill = white_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if column != 2 else "left", vertical="center")
        ws.cell(row=row, column=3).number_format = "0.00000"
        ws.cell(row=row, column=4).number_format = "0%"
        ws.cell(row=row, column=7).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=8).number_format = '"$"#,##0.00'

        for column in (5, 6):
            set_cell_value(ws, row, column, None)
            ws.cell(row=row, column=column).fill = white_fill

    total_row = header_row + 4
    clear_cells(ws, [total_row], range(1, 9))
    set_cell_value(ws, total_row, 7, "Total Price\n(USD/1,000 Sticks)")
    set_cell_value(ws, total_row, 8, f"=SUM(H{data_rows[0]}:H{data_rows[-1]})")
    for column in (7, 8):
        cell = ws.cell(row=total_row, column=column)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.number_format = "General"
    ws.cell(row=total_row, column=8).number_format = '"$"#,##0.00'


def strip_xlsx_repair_risk_artifacts(workbook: openpyxl.Workbook) -> None:
    workbook._external_links = []
    for worksheet in workbook.worksheets:
        worksheet._images = []
        worksheet._charts = []


def apply_basic_sheet_style(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    ws.sheet_view.showGridLines = True
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = copy(cell.alignment)


def place_logo_in_merged_cell(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    logo_path: Path = LOGO_PATH,
    merged_range: str = LOGO_MERGED_RANGE,
) -> None:
    existing_ranges = {str(cell_range) for cell_range in ws.merged_cells.ranges}
    if merged_range not in existing_ranges:
        ws.merge_cells(merged_range)

    logo_cell = ws[merged_range.split(":")[0]]
    logo_cell.value = None
    logo_cell.alignment = Alignment(horizontal="center", vertical="center")

    if not logo_path.exists():
        return

    try:
        logo = ExcelImage(str(logo_path))
        logo.width = 185
        logo.height = 48
        ws.add_image(logo, logo_cell.coordinate)
    except Exception:
        pass


def setup_clean_formula_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, logo_path: Path = LOGO_PATH) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    table_fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for column in range(1, 18):
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 16
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["Q"].width = 38

    place_logo_in_merged_cell(ws, logo_path)
    ws.merge_cells("C1:O2")
    ws["C1"] = "Laboratory Formulation Sheet\n& Lab Work Instruction"
    ws["C1"].font = Font(bold=True, size=16)
    ws["C1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["P1"] = "Document No"
    ws["Q1"] = "BOL-ID-BOLL-SPEC-003"
    ws["P2"] = "Revision No"
    ws["Q2"] = "0"
    ws["P3"] = "Effective Date"

    labels = {
        "A4": "Factory Lab to Fill Up",
        "A5": "Full FG Description",
        "A6": "Mixing Factory",
        "C6": "PT SAAT",
        "A7": "Lab Personnel Involved",
        "A9": "Flavor Development Reference",
        "A10": "Standard Control",
        "H10": "Flavor Standard Reference",
        "A12": "Product Specification",
        "A13": "Formulation Code",
        "H12": "Sensory Parameter",
        "H13": "Impact",
        "A14": "Single Capsule",
        "H14": "Flavor Aroma",
        "A15": "Double Capsule (Tobacco End)",
        "H15": "Irritation",
        "A16": "Double Capsule (Mouth End)",
        "H16": "Cooling",
        "A18": "Product Weight (mg/stick)",
        "E18": "Clove Weight (mg/stick)",
        "I18": "Stick per MC",
    }
    for cell_ref, value in labels.items():
        ws[cell_ref] = value
        ws[cell_ref].font = Font(bold=True)
        ws[cell_ref].fill = section_fill

    for cell in ws["P1:Q3"][0] + ws["P1:Q3"][1] + ws["P1:Q3"][2]:
        cell.border = border
    for row in range(4, 19):
        for column in range(1, 18):
            ws.cell(row=row, column=column).border = border

    material_headers = [
        "No",
        "Material NAV Item\nCode",
        "Material Name",
        "",
        "Physical\nForm",
        "CAS Number",
        "Ratio\n(%)",
        "Dosage\n(mg/stick)",
        "Material\nPrice\n(USD / KG)",
        "Formulation\nPrice\n(USD / KG)",
        "Dosage\n(KG / MC)",
        "Density",
        "Addition\nSequence",
        "Temperature",
        "Agitation\nRate",
        "Mixing\nDuration",
        "Work Instruction",
    ]

    phase_titles = {
        "Casing Rajangan": (20, "Casing", True),
        "Casing Krosok": (33, "Casing", True),
        "Top Flavor": (46, "Top Flavor", True),
        "Casing Pre-Mix": (73, "Casing Pre-Mix Formulation", False),
        "Flavor Pre-Mix 1": (89, "Flavor Pre-Mix Formulation", False),
        "Flavor Pre-Mix 2": (130, "Flavor Pre-Mix Formulation", False),
    }
    for phase, (title_row, label_prefix, has_ratio) in phase_titles.items():
        positions = PHASE_METADATA_POSITIONS[phase]
        ws.cell(row=title_row, column=1, value=phase if phase in STATIC_PHASES else label_prefix)
        ws.cell(row=title_row, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=title_row, column=1).fill = header_fill
        ws.cell(row=positions["nav_code"][0], column=1, value=f"{label_prefix} NAV Item Code")
        ws.cell(row=positions["description"][0], column=1, value=f"{label_prefix} NAV Item Description")
        if has_ratio:
            ws.cell(row=positions["blend_ratio"][0], column=1, value="Blend Ratio")
            ws.cell(row=positions["application"][0], column=1, value=f"{label_prefix} Application (%)")

        header_row = SECTION_ROW_RANGES[phase][0] - 1
        for column, value in enumerate(material_headers, start=1):
            cell = ws.cell(row=header_row, column=column, value=value)
            cell.font = Font(bold=True)
            cell.fill = table_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in SECTION_ROW_RANGES[phase] + [SECTION_ROW_RANGES[phase][-1] + 1]:
            for column in range(1, 18):
                cell = ws.cell(row=row, column=column)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 22
    ws.freeze_panes = "A25"


def clone_template_sheet_layout(
    source_ws: openpyxl.worksheet.worksheet.Worksheet,
    target_ws: openpyxl.worksheet.worksheet.Worksheet,
    logo_path: Path = LOGO_PATH,
) -> None:
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_format = copy(source_ws.sheet_format)
    target_ws.sheet_properties = copy(source_ws.sheet_properties)
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.print_options = copy(source_ws.print_options)

    for column_letter, dimension in source_ws.column_dimensions.items():
        target_dimension = target_ws.column_dimensions[column_letter]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden
        target_dimension.outlineLevel = dimension.outlineLevel

    for row_index, dimension in source_ws.row_dimensions.items():
        target_dimension = target_ws.row_dimensions[row_index]
        target_dimension.height = dimension.height
        target_dimension.hidden = dimension.hidden
        target_dimension.outlineLevel = dimension.outlineLevel

    for row in source_ws.iter_rows():
        for source_cell in row:
            if isinstance(source_cell, MergedCell):
                continue
            target_cell = target_ws.cell(
                row=source_cell.row,
                column=source_cell.column,
                value=source_cell.value,
            )
            copy_cell_format(source_cell, target_cell)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    if source_ws.auto_filter.ref:
        target_ws.auto_filter.ref = source_ws.auto_filter.ref

    place_logo_in_merged_cell(target_ws, logo_path)


def write_material_db_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, template_path: Path) -> None:
    df = pd.read_excel(template_path, sheet_name="BOL-SAAT List", engine="openpyxl")
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for column_index, column_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=column_index, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
    for row_index, row in enumerate(df.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.border = border
    for column in range(1, len(df.columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 22
    ws.freeze_panes = "A2"


def write_input_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, formulation: FormulationInput) -> None:
    for field_name, position in INPUT_METADATA_POSITIONS.items():
        value = getattr(formulation, field_name)
        ws.cell(row=position[0], column=position[1], value=value)

    clear_until = max(ws.max_row, MATERIAL_START_ROW + len(formulation.materials) + 50)
    clear_cells(ws, range(MATERIAL_START_ROW, clear_until + 1), MATERIAL_COLUMNS.values())

    for offset, material in enumerate(formulation.materials, start=0):
        row_index = MATERIAL_START_ROW + offset
        for field_name, column in MATERIAL_COLUMNS.items():
            value = getattr(material, field_name)
            if value is None:
                continue
            if field_name == "dosage_mg_stick":
                ws.cell(row=row_index, column=column, value=float(value))
            elif field_name in ("ratio_percent", "application_percent"):
                ws.cell(row=row_index, column=column, value=float(value))
            else:
                ws.cell(row=row_index, column=column, value=value)


def write_formula_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, formulation: FormulationInput) -> None:
    phase_rows: Dict[str, List[MaterialInput]] = {}
    for material in formulation.materials:
        phase_rows.setdefault(material.phase, []).append(material)

    ensure_dynamic_premix_layouts(ws, phase_rows)
    prepare_dynamic_formula_sections(ws, phase_rows)

    for field_name, position in FORMULA_METADATA_POSITIONS.items():
        value = getattr(formulation, field_name)
        ws.cell(row=position[0], column=position[1], value=value)

    apply_document_header_layout(ws, formulation)

    if formulation.phase_metadata:
        for phase, metadata in formulation.phase_metadata.items():
            phase_key = normalize_phase(phase)
            positions = PHASE_METADATA_POSITIONS.get(phase_key)
            if not positions:
                continue
            if metadata.nav_code and positions.get("nav_code"):
                ws.cell(row=positions["nav_code"][0], column=positions["nav_code"][1], value=metadata.nav_code)
            if metadata.description and positions.get("description"):
                ws.cell(row=positions["description"][0], column=positions["description"][1], value=metadata.description)
            if metadata.blend_ratio is not None and positions.get("blend_ratio"):
                ws.cell(row=positions["blend_ratio"][0], column=positions["blend_ratio"][1], value=metadata.blend_ratio)
            if metadata.application is not None and positions.get("application"):
                ws.cell(row=positions["application"][0], column=positions["application"][1], value=metadata.application)

    ws.cell(
        row=FORMULA_METADATA_POSITIONS["formulation_code"][0],
        column=FORMULA_METADATA_POSITIONS["formulation_code"][1],
        value=build_formulation_code_formula(),
    )

    clear_cells(
        ws,
        [row for rows in SECTION_ROW_RANGES.values() for row in rows],
        FORMULA_MATERIAL_COLUMNS.values(),
    )

    for phase, rows in phase_rows.items():
        if phase not in SECTION_ROW_RANGES:
            raise ValueError(f"Unsupported phase: {phase}. Gunakan salah satu phase: {', '.join(KNOWN_PHASES)}")
        row_indices = SECTION_ROW_RANGES[phase]
        for row_index, material in zip(row_indices, rows):
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["no"], row_indices.index(row_index) + 1)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["item_code"], material.item_code)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["item_name"], material.item_name)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["physical_form"], material.physical_form)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["cas_number"], material.cas_number)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["ratio"], material.ratio)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"], material.dosage_mg_stick)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["material_price"], material.material_price)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["formulation_price"], material.formulation_price)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["dosage_kg_mc"], material.dosage_kg_mc)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["density"], None)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["addition_sequence"], material.addition_sequence)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["temperature"], material.temperature)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["agitation_rate"], material.agitation_rate)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["mixing_duration"], material.mixing_duration)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["work_instruction"], material.work_instruction)

    regenerate_total_rows(ws)
    apply_formula_number_formats(ws)
    merge_process_instruction_columns(ws, phase_rows)
    update_top_formula_summary(ws)
    apply_approval_block(ws, formulation)


def ensure_xlsx_output_path(output_path: Path) -> Path:
    if output_path.suffix.lower() != ".xlsx":
        return output_path.with_suffix(".xlsx")
    return output_path


def generate_formulation_workbook(
    formulation: FormulationInput,
    template_path: Path,
    output_path: Path,
) -> Path:
    output_path = ensure_xlsx_output_path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    reset_runtime_layout()
    material_db = load_material_db(template_path)
    validation_errors = validate_formulation_input(formulation, material_db)
    if validation_errors:
        raise ValueError("Validasi formulasi gagal:\n- " + "\n- ".join(validation_errors))
    compute_materials(formulation, material_db)

    template_workbook = openpyxl.load_workbook(template_path, keep_vba=False, keep_links=False)
    if "Formula" not in template_workbook.sheetnames:
        raise ValueError("Template workbook harus memiliki sheet 'Formula'.")

    workbook = openpyxl.load_workbook(BytesIO(create_blank_input_workbook(template_path)))
    formula_sheet = workbook.create_sheet("Formula", 0)
    material_db_sheet = workbook.create_sheet("BOL-SAAT List", 1)
    clone_template_sheet_layout(template_workbook["Formula"], formula_sheet)
    write_material_db_sheet(material_db_sheet, template_path)

    write_input_sheet(workbook["INPUT"], formulation)
    write_formula_sheet(workbook["Formula"], formulation)

    workbook.save(output_path)
    return output_path


def create_blank_input_workbook(template_path: Path) -> bytes:
    workbook = openpyxl.Workbook()
    input_sheet = workbook.active
    input_sheet.title = "INPUT"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    table_fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    input_sheet.merge_cells("A1:N1")
    input_sheet["A1"] = "FORMULATION INPUT"
    input_sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    input_sheet["A1"].fill = header_fill
    input_sheet["A1"].alignment = Alignment(horizontal="center")

    input_sheet.merge_cells("D2:H2")
    input_sheet["D2"] = "PHASE / CASING / FLAVOR METADATA"
    input_sheet.merge_cells("J2:N2")
    input_sheet["J2"] = "FLAVOR DEVELOPMENT REFERENCE"
    input_sheet.merge_cells("J4:N4")
    input_sheet["J4"] = "PRODUCT SPECIFICATION"

    for cell_ref in ("D2", "J2", "J4"):
        cell = input_sheet[cell_ref]
        cell.font = Font(bold=True)
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")

    metadata_labels = {
        (2, 1): "Product Name",
        (3, 1): "Formula Code",
        (4, 1): "Product Weight (mg/stick)",
        (5, 1): "Clove Weight (mg/stick)",
        (6, 1): "Stick per MC",
        (7, 1): "Prepared By",
        (8, 1): "Date",
        (3, 10): "Standard Control",
        (3, 13): "Flavor Standard Reference",
        (5, 10): "Tobacco Blend Code",
        (5, 13): "Sensory Parameter",
        (6, 10): "Formulation Code",
        (6, 13): "Impact",
        (7, 10): "Single Capsule",
        (7, 13): "Flavor Aroma",
        (8, 10): "Double Capsule (Tobacco End)",
        (8, 13): "Irritation",
        (9, 10): "Double Capsule (Mouth End)",
        (9, 13): "Cooling",
    }
    for (row, column), label in metadata_labels.items():
        cell = input_sheet.cell(row=row, column=column, value=label)
        cell.font = Font(bold=True)
        cell.fill = section_fill
        cell.border = border

    for row in range(2, 9):
        input_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    for row in (3, 5, 6, 7, 8, 9):
        input_sheet.merge_cells(start_row=row, start_column=11, end_row=row, end_column=12)

    phase_headers = ["Phase", "NAV Item Code", "NAV Item Description", "Blend Ratio", "Application %"]
    for offset, header in enumerate(phase_headers, start=4):
        cell = input_sheet.cell(row=3, column=offset, value=header)
        cell.font = Font(bold=True)
        cell.fill = table_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    metadata_phases = MAIN_PHASES + [
        "Casing Pre-Mix",
        "Casing Pre-Mix 2",
        "Casing Pre-Mix 3",
        "Flavor Pre-Mix 1",
        "Flavor Pre-Mix 2",
        "Flavor Pre-Mix 3",
        "Flavor Pre-Mix 4",
        "Flavor Pre-Mix 5",
    ]
    for row_offset, phase in enumerate(metadata_phases, start=PHASE_METADATA_START_ROW):
        input_sheet.cell(row=row_offset, column=4, value=phase)
        for column in range(4, 9):
            input_sheet.cell(row=row_offset, column=column).border = border

    material_headers = {
        "phase": "Phase",
        "item_code": "Item Code",
        "item_name": "Item Name",
        "dosage_input_mode": "Input Mode",
        "dosage_mg_stick": "Dosage mg/stick",
        "ratio_percent": "Ratio %",
        "addition_sequence": "Addition Sequence",
        "temperature": "Temperature",
        "agitation_rate": "Agitation Rate",
        "mixing_duration": "Mixing Duration",
        "work_instruction_override": "Work Instruction Override",
        "process_role": "Process Role",
        "notes": "Notes",
    }
    for field_name, column in MATERIAL_COLUMNS.items():
        cell = input_sheet.cell(row=MATERIAL_HEADER_ROW, column=column, value=material_headers[field_name])
        cell.font = Font(bold=True)
        cell.fill = table_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row in range(MATERIAL_START_ROW, MATERIAL_START_ROW + 250):
        for column in MATERIAL_COLUMNS.values():
            input_sheet.cell(row=row, column=column).border = border
        input_sheet.cell(row=row, column=MATERIAL_COLUMNS["dosage_mg_stick"]).number_format = "0.00000"
        input_sheet.cell(row=row, column=MATERIAL_COLUMNS["ratio_percent"]).number_format = "0.00000"

    widths = {
        "A": 24,
        "B": 18,
        "C": 30,
        "D": 18,
        "E": 20,
        "F": 24,
        "G": 14,
        "H": 16,
        "I": 22,
        "J": 24,
        "K": 24,
        "L": 22,
        "M": 24,
        "N": 24,
    }
    for column, width in widths.items():
        input_sheet.column_dimensions[column].width = width
    input_sheet.freeze_panes = f"A{MATERIAL_START_ROW}"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_formulation_input_from_dict(data: Dict[str, Any], material_rows: Iterable[Dict[str, Any]]) -> FormulationInput:
    phase_metadata: Dict[str, PhaseMetadata] = {}
    for phase_name, values in (data.get("phase_metadata") or {}).items():
        if phase_name:
            phase_metadata[normalize_phase(phase_name)] = PhaseMetadata(
                phase=phase_name,
                nav_code=values.get("nav_code"),
                description=values.get("description"),
                blend_ratio=values.get("blend_ratio"),
                application=values.get("application"),
            )

    materials = []
    for row in material_rows:
        if not row.get("phase") or pd.isna(row.get("phase")):
            continue
        dosage = row.get("dosage_mg_stick")
        dosage_input_mode = row.get("dosage_input_mode") or "mg/stick"
        if normalize_input_mode(dosage_input_mode) != "%" and (dosage is None or dosage == ""):
            continue
        materials.append(MaterialInput(
            phase=normalize_phase(row.get("phase")),
            item_code=row.get("item_code"),
            item_name=row.get("item_name"),
            dosage_mg_stick=float(dosage) if dosage not in (None, "") else None,
            dosage_input_mode=dosage_input_mode,
            ratio_percent=row.get("ratio_percent"),
            application_percent=row.get("application_percent"),
            addition_sequence=int(row.get("addition_sequence")) if row.get("addition_sequence") not in (None, "") else None,
            temperature=row.get("temperature"),
            agitation_rate=row.get("agitation_rate"),
            mixing_duration=row.get("mixing_duration"),
            work_instruction_override=row.get("work_instruction_override"),
            process_role=row.get("process_role"),
            notes=row.get("notes"),
        ))

    return FormulationInput(
        product_name=data.get("product_name", ""),
        formula_code=data.get("formula_code", ""),
        product_weight_mg_stick=float(data.get("product_weight_mg_stick", 0)),
        clove_weight_mg_stick=float(data.get("clove_weight_mg_stick", 0)),
        stick_per_mc=int(data.get("stick_per_mc", 10000)),
        prepared_by=data.get("prepared_by", ""),
        date=data.get("date"),
        standard_control=data.get("standard_control", ""),
        flavor_standard_reference=data.get("flavor_standard_reference", ""),
        tobacco_blend_code=data.get("tobacco_blend_code", ""),
        sensory_parameter=data.get("sensory_parameter", ""),
        formulation_code=data.get("formulation_code", ""),
        impact=data.get("impact", ""),
        single_capsule=data.get("single_capsule", ""),
        flavor_aroma=data.get("flavor_aroma", ""),
        double_capsule_tobacco_end=data.get("double_capsule_tobacco_end", ""),
        irritation=data.get("irritation", ""),
        double_capsule_mouth_end=data.get("double_capsule_mouth_end", ""),
        cooling=data.get("cooling", ""),
        phase_metadata=phase_metadata,
        materials=materials,
    )


def _serialize_model_value(value: Any) -> Any:
    if isinstance(value, (datetime, date_class)):
        return value.isoformat()
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    return value


def export_formulation_model(formulation: FormulationInput) -> Dict[str, Any]:
    product = {
        "product_name": formulation.product_name,
        "formula_code": formulation.formula_code,
        "product_weight_mg_stick": formulation.product_weight_mg_stick,
        "clove_weight_mg_stick": formulation.clove_weight_mg_stick,
        "stick_per_mc": formulation.stick_per_mc,
        "prepared_by": formulation.prepared_by,
        "date": _serialize_model_value(formulation.date),
        "standard_control": formulation.standard_control,
        "flavor_standard_reference": formulation.flavor_standard_reference,
        "tobacco_blend_code": formulation.tobacco_blend_code,
        "sensory_parameter": formulation.sensory_parameter,
        "formulation_code": formulation.formulation_code,
        "impact": formulation.impact,
        "single_capsule": formulation.single_capsule,
        "flavor_aroma": formulation.flavor_aroma,
        "double_capsule_tobacco_end": formulation.double_capsule_tobacco_end,
        "irritation": formulation.irritation,
        "double_capsule_mouth_end": formulation.double_capsule_mouth_end,
        "cooling": formulation.cooling,
    }

    phase_metadata = [
        {
            "phase": metadata.phase,
            "nav_code": metadata.nav_code,
            "description": metadata.description,
            "blend_ratio": metadata.blend_ratio,
            "application": metadata.application,
        }
        for metadata in sorted(formulation.phase_metadata.values(), key=lambda item: phase_sort_key(item.phase))
    ]

    materials = [
        {
            "phase": material.phase,
            "item_code": material.item_code,
            "item_name": material.item_name,
            "dosage_input_mode": material.dosage_input_mode,
            "dosage_mg_stick": material.dosage_mg_stick,
            "ratio_percent": material.ratio_percent,
            "addition_sequence": material.addition_sequence,
            "temperature": material.temperature,
            "agitation_rate": material.agitation_rate,
            "mixing_duration": material.mixing_duration,
            "work_instruction_override": material.work_instruction_override,
            "process_role": material.process_role,
            "notes": material.notes,
        }
        for material in sorted(
            formulation.materials,
            key=lambda item: (
                phase_sort_key(item.phase),
                item.addition_sequence if item.addition_sequence is not None else 999999,
                item.item_code or "",
            ),
        )
    ]

    return {
        "schema_version": "1.0",
        "model_type": "formulation_product",
        "product": product,
        "phase_metadata": phase_metadata,
        "materials": materials,
    }


def export_formulation_model_json(formulation: FormulationInput, *, indent: int = 2) -> bytes:
    model = export_formulation_model(formulation)
    return json.dumps(model, ensure_ascii=False, indent=indent, default=str).encode("utf-8")


def _dicts_to_csv_bytes(rows: List[Dict[str, Any]], fieldnames: List[str]) -> bytes:
    text_io = io.StringIO()
    writer = csv.DictWriter(text_io, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({field: _serialize_model_value(row.get(field)) for field in fieldnames})
    return text_io.getvalue().encode("utf-8")


def export_formulation_model_csv_bundle(formulation: FormulationInput) -> bytes:
    model = export_formulation_model(formulation)
    product_rows = [model["product"]]
    phase_rows = model["phase_metadata"]
    material_rows = model["materials"]

    csv_files = {
        "product.csv": _dicts_to_csv_bytes(product_rows, list(model["product"].keys())),
        "phase_metadata.csv": _dicts_to_csv_bytes(
            phase_rows,
            ["phase", "nav_code", "description", "blend_ratio", "application"],
        ),
        "materials.csv": _dicts_to_csv_bytes(
            material_rows,
            [
                "phase",
                "item_code",
                "item_name",
                "dosage_input_mode",
                "dosage_mg_stick",
                "ratio_percent",
                "addition_sequence",
                "temperature",
                "agitation_rate",
                "mixing_duration",
                "work_instruction_override",
                "process_role",
                "notes",
            ],
        ),
    }

    buffer = BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as archive:
        for filename, content in csv_files.items():
            archive.writestr(filename, content)
    buffer.seek(0)
    return buffer.getvalue()
