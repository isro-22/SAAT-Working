from datetime import date as date_class, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
import streamlit as st
from formulation_generator import (
    INPUT_METADATA_POSITIONS,
    MATERIAL_COLUMNS,
    MATERIAL_HEADER_ROW,
    MATERIAL_START_ROW,
    PHASE_METADATA_ROW_COUNT,
    PHASE_METADATA_START_ROW,
    build_formulation_input_from_dict,
    create_blank_input_workbook,
    export_formulation_model_csv_bundle,
    export_formulation_model_json,
    generate_formulation_workbook,
)

PROJECT_ROOT = Path(__file__).resolve().parent
TEMPLATE_DIR = PROJECT_ROOT / "templates"
OUTPUT_DIR = PROJECT_ROOT / "outputs" / "generated"
TEMPLATE_PATH = TEMPLATE_DIR / "Template_Generate.xlsm"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

MAIN_PHASES = [
    "Casing Rajangan",
    "Casing Krosok",
    "Top Flavor",
]

DEFAULT_PHASES = MAIN_PHASES + [
    "Casing Pre-Mix",
    "Flavor Pre-Mix 1",
    "Flavor Pre-Mix 2",
]


def build_phase_options(casing_premix_count: int = 1, flavor_premix_count: int = 2) -> List[str]:
    phases = MAIN_PHASES.copy()
    phases.extend(
        "Casing Pre-Mix" if index == 1 else f"Casing Pre-Mix {index}"
        for index in range(1, casing_premix_count + 1)
    )
    phases.extend(f"Flavor Pre-Mix {index}" for index in range(1, flavor_premix_count + 1))
    return phases


def _normalize_uploaded_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _parse_optional_float(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _parse_optional_int(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    if isinstance(value, float) and not value.is_integer():
        return None
    try:
        parsed_value = int(value)
    except (TypeError, ValueError):
        return None
    if isinstance(value, str) and str(parsed_value) != value.strip():
        return None
    return parsed_value


def _coerce_date_value(value: Any) -> date_class:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date_class):
        return value
    if isinstance(value, str) and value.strip():
        for date_format in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value.strip(), date_format).date()
            except ValueError:
                continue
    return date_class.today()


def _is_empty(value: Any) -> bool:
    if value in (None, ""):
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _row_is_blank(row_values: Dict[str, Any]) -> bool:
    return all(_is_empty(value) for value in row_values.values())


def _phase_application_value(phase_metadata: Dict[str, Dict[str, Any]], phase: Any) -> Any:
    if _is_empty(phase):
        return None
    return (phase_metadata.get(str(phase), {}) or {}).get("application")


def _validate_materials(material_rows: List[Dict[str, Any]], phase_metadata: Dict[str, Dict[str, Any]]) -> List[str]:
    errors: List[str] = []
    valid_rows = 0
    for index, row in enumerate(material_rows, start=1):
        phase = row.get("phase")
        input_mode = str(row.get("dosage_input_mode") or "mg/stick").strip()
        dosage = row.get("dosage_mg_stick")
        ratio_percent = row.get("ratio_percent")
        application_percent = _phase_application_value(phase_metadata, phase)
        item_code = row.get("item_code")
        item_name = row.get("item_name")

        if _row_is_blank(row):
            continue

        row_label = f"Material row {index}"
        if _is_empty(phase):
            errors.append(f"{row_label}: phase wajib diisi.")
        if _is_empty(item_code) and _is_empty(item_name):
            errors.append(f"{row_label}: item code atau item name wajib diisi.")
        if input_mode == "%":
            try:
                ratio_value = float(ratio_percent)
                application_value = float(application_percent)
                if ratio_value <= 0 or application_value <= 0:
                    errors.append(f"{row_label}: Ratio % dan Application % harus lebih dari 0.")
            except (TypeError, ValueError):
                errors.append(f"{row_label}: Ratio % wajib angka dan Application % wajib diisi di Phase Metadata untuk mode %.")
            else:
                if not _is_empty(phase):
                    valid_rows += 1
        else:
            try:
                dosage_value = float(dosage)
                if dosage_value <= 0:
                    errors.append(f"{row_label}: dosage harus lebih dari 0.")
            except (TypeError, ValueError):
                errors.append(f"{row_label}: dosage harus berupa angka untuk mode mg/stick.")
            else:
                if dosage_value > 0 and not _is_empty(phase):
                    valid_rows += 1

        addition_sequence = row.get("addition_sequence")
        if not _is_empty(addition_sequence):
            try:
                parsed_sequence = int(addition_sequence)
            except (TypeError, ValueError):
                errors.append(f"{row_label}: addition sequence harus bilangan bulat.")
            else:
                if isinstance(addition_sequence, float) and not addition_sequence.is_integer():
                    errors.append(f"{row_label}: addition sequence harus bilangan bulat.")
                if isinstance(addition_sequence, str) and str(parsed_sequence) != addition_sequence.strip():
                    errors.append(f"{row_label}: addition sequence harus bilangan bulat.")

    if valid_rows == 0:
        errors.append("Minimal satu material valid wajib diisi.")
    return errors


def collect_material_issues(material_rows: List[Dict[str, Any]], phase_metadata: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    valid_rows = 0
    phases_with_valid_rows = set()

    def add_issue(row_number: Any, field: str, message: str, severity: str = "Error") -> None:
        issues.append({
            "Severity": severity,
            "Row": row_number,
            "Field": field,
            "Message": message,
        })

    for index, row in enumerate(material_rows, start=1):
        if _row_is_blank(row):
            continue

        phase = row.get("phase")
        input_mode = str(row.get("dosage_input_mode") or "mg/stick").strip()
        dosage = row.get("dosage_mg_stick")
        ratio_percent = row.get("ratio_percent")
        application_percent = _phase_application_value(phase_metadata, phase)
        item_code = row.get("item_code")
        item_name = row.get("item_name")
        row_valid = True

        if _is_empty(phase):
            add_issue(index, "phase", "Phase wajib diisi.")
            row_valid = False
        if _is_empty(item_code) and _is_empty(item_name):
            add_issue(index, "item_code / item_name", "Isi minimal salah satu: item_code atau item_name.")
            row_valid = False
        if input_mode not in ("mg/stick", "%"):
            add_issue(index, "Input Mode", "Pilih input mode: mg/stick atau %.")
            row_valid = False
        elif input_mode == "%":
            if _is_empty(ratio_percent):
                add_issue(index, "Ratio %", "Wajib diisi untuk mode %.")
                row_valid = False
            if _is_empty(application_percent):
                add_issue(index, "Phase Metadata Application %", f"Isi Application % pada Phase Metadata untuk phase {phase}.")
                row_valid = False
            if not _is_empty(ratio_percent):
                try:
                    if float(ratio_percent) <= 0:
                        add_issue(index, "Ratio %", "Harus lebih dari 0.")
                        row_valid = False
                except (TypeError, ValueError):
                    add_issue(index, "Ratio %", "Harus berupa angka.")
                    row_valid = False
            if not _is_empty(application_percent):
                try:
                    if float(application_percent) <= 0:
                        add_issue(index, "Phase Metadata Application %", "Harus lebih dari 0.")
                        row_valid = False
                except (TypeError, ValueError):
                    add_issue(index, "Phase Metadata Application %", "Harus berupa angka.")
                    row_valid = False
        else:
            if _is_empty(dosage):
                add_issue(index, "Dosage (mg/stick)", "Wajib diisi untuk mode mg/stick.")
                row_valid = False
            else:
                try:
                    if float(dosage) <= 0:
                        add_issue(index, "Dosage (mg/stick)", "Harus lebih dari 0.")
                        row_valid = False
                except (TypeError, ValueError):
                    add_issue(index, "Dosage (mg/stick)", "Harus berupa angka.")
                    row_valid = False

        addition_sequence = row.get("addition_sequence")
        if not _is_empty(addition_sequence):
            try:
                parsed_sequence = int(addition_sequence)
            except (TypeError, ValueError):
                add_issue(index, "Addition Sequence", "Harus bilangan bulat.")
                row_valid = False
            else:
                if isinstance(addition_sequence, float) and not addition_sequence.is_integer():
                    add_issue(index, "Addition Sequence", "Harus bilangan bulat.")
                    row_valid = False
                if isinstance(addition_sequence, str) and str(parsed_sequence) != addition_sequence.strip():
                    add_issue(index, "Addition Sequence", "Harus bilangan bulat.")
                    row_valid = False

        if row_valid:
            valid_rows += 1
            phases_with_valid_rows.add(phase)

    if valid_rows == 0:
        add_issue("-", "Material Table", "Minimal satu material valid wajib diisi.")
    for phase in MAIN_PHASES:
        if phase not in phases_with_valid_rows:
            add_issue("-", "Phase utama", f"{phase} wajib memiliki minimal satu material valid.")
    return issues


def collect_header_issues(
    product_name: str,
    formula_code: str,
    prepared_by: str,
    product_weight_mg_stick: float,
    clove_weight_mg_stick: float,
    stick_per_mc: int,
) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []

    def add_issue(field: str, message: str) -> None:
        issues.append({"Severity": "Error", "Row": "-", "Field": field, "Message": message})

    if not product_name.strip():
        add_issue("Product Name", "Wajib diisi.")
    if not formula_code.strip():
        add_issue("Formula Code", "Wajib diisi.")
    if not prepared_by.strip():
        add_issue("Prepared By", "Wajib diisi.")
    if product_weight_mg_stick <= 0:
        add_issue("Product Weight", "Harus lebih dari 0.")
    if clove_weight_mg_stick < 0:
        add_issue("Clove Weight", "Tidak boleh negatif.")
    if stick_per_mc <= 0:
        add_issue("Stick per MC", "Harus lebih dari 0.")
    return issues


def parse_excel_input(uploaded_file) -> Tuple[Dict[str, Any], List[Dict[str, Any]], List[str]]:
    uploaded_file.seek(0)
    workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
    if "INPUT" not in workbook.sheetnames:
        raise ValueError("Sheet 'INPUT' tidak ditemukan dalam file Excel.")

    ws = workbook["INPUT"]
    data: Dict[str, Any] = {}
    for field_name, position in INPUT_METADATA_POSITIONS.items():
        data[field_name] = _normalize_uploaded_value(ws.cell(row=position[0], column=position[1]).value)

    phase_metadata: Dict[str, Dict[str, Any]] = {}
    for row in range(PHASE_METADATA_START_ROW, PHASE_METADATA_START_ROW + PHASE_METADATA_ROW_COUNT):
        phase_name = _normalize_uploaded_value(ws.cell(row=row, column=4).value)
        if not phase_name:
            continue
        nav_code = _normalize_uploaded_value(ws.cell(row=row, column=5).value)
        description = _normalize_uploaded_value(ws.cell(row=row, column=6).value)
        blend_ratio = _parse_optional_float(ws.cell(row=row, column=7).value)
        application = _parse_optional_float(ws.cell(row=row, column=8).value)
        if nav_code or description or blend_ratio is not None or application is not None:
            phase_metadata[phase_name] = {
                "nav_code": nav_code,
                "description": description,
                "blend_ratio": blend_ratio,
                "application": application,
            }
    data["phase_metadata"] = phase_metadata

    materials: List[Dict[str, Any]] = []
    warnings: List[str] = []
    for row in range(MATERIAL_START_ROW, MATERIAL_START_ROW + 200):
        row_values = {
            field_name: _normalize_uploaded_value(ws.cell(row=row, column=column).value)
            for field_name, column in MATERIAL_COLUMNS.items()
        }
        if _row_is_blank(row_values):
            continue
        input_mode = row_values.get("dosage_input_mode") or "mg/stick"
        if not row_values["phase"]:
            warnings.append(f"Baris Excel {row}: dilewati karena phase kosong.")
            continue
        dosage = row_values["dosage_mg_stick"]
        if input_mode == "%":
            dosage = None
        else:
            if not dosage:
                warnings.append(f"Baris Excel {row}: dilewati karena dosage kosong.")
                continue
            try:
                dosage = float(dosage)
            except (TypeError, ValueError):
                warnings.append(f"Baris Excel {row}: dilewati karena dosage bukan angka ({dosage}).")
                continue
        addition_sequence = _parse_optional_int(row_values["addition_sequence"])
        if row_values["addition_sequence"] not in (None, "") and addition_sequence is None:
            warnings.append(f"Baris Excel {row}: addition sequence bukan bilangan bulat dan dikosongkan.")
        materials.append({
            "phase": row_values["phase"],
            "item_code": row_values["item_code"],
            "item_name": row_values["item_name"],
            "dosage_input_mode": input_mode,
            "dosage_mg_stick": dosage,
            "ratio_percent": _parse_optional_float(row_values.get("ratio_percent")),
            "addition_sequence": addition_sequence,
            "temperature": row_values["temperature"],
            "agitation_rate": row_values["agitation_rate"],
            "mixing_duration": row_values["mixing_duration"],
            "work_instruction_override": row_values["work_instruction_override"],
            "process_role": row_values["process_role"],
            "notes": row_values["notes"],
        })

    return data, materials, warnings

st.set_page_config(page_title="Formulation Sheet Generator", layout="wide")
st.title("Formulation Sheet Generator")
st.markdown(
    "Masukkan data formula dan material. Anda bisa upload file Excel `INPUT`, lalu review dan generate formulasi dari Streamlit."
)

if TEMPLATE_PATH.exists():
    try:
        blank_input_bytes = create_blank_input_workbook(TEMPLATE_PATH)
        st.download_button(
            "Download blank INPUT template",
            blank_input_bytes,
            file_name="Blank_Formulation_INPUT.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        st.warning(f"Blank INPUT template belum bisa dibuat: {exc}")

uploaded_input = st.file_uploader("Upload file Excel INPUT", type=["xlsx", "xlsm"])
excel_data: Dict[str, Any] = {}
excel_materials: List[Dict[str, Any]] = []
upload_warnings: List[str] = []
upload_error = None
if uploaded_input:
    try:
        excel_data, excel_materials, upload_warnings = parse_excel_input(uploaded_input)
        st.success("Data INPUT berhasil dimuat dari file. Silakan review dan sesuaikan jika perlu.")
        if upload_warnings:
            with st.expander(f"Upload warnings ({len(upload_warnings)})", expanded=True):
                for warning in upload_warnings:
                    st.warning(warning)
        with st.expander("Preview INPUT loaded from Excel", expanded=True):
            preview_data = {k: v for k, v in excel_data.items() if k != "phase_metadata"}
            st.write("**Header / metadata values**")
            st.json(preview_data)
            st.write("**Phase metadata**")
            st.json(excel_data.get("phase_metadata", {}))
            if excel_materials:
                st.write("**Materials loaded from INPUT**")
                st.dataframe(pd.DataFrame(excel_materials), width="stretch")
    except Exception as exc:
        upload_error = str(exc)
        st.error(upload_error)

with st.expander("Product & Flavor Metadata", expanded=True):
    col1, col2, col3 = st.columns(3)
    with col1:
        product_name = st.text_input("Product Name", value=excel_data.get("product_name", "SAAT 369"))
        formula_code = st.text_input("Formula Code", value=excel_data.get("formula_code", "CS_K-CK-LCDT-04_R00"))
        product_weight_mg_stick = st.number_input(
            "Product Weight (mg/stick)",
            value=float(excel_data.get("product_weight_mg_stick", 720.0)) if excel_data.get("product_weight_mg_stick") not in (None, "") else 720.0,
            step=1.0,
        )
        clove_weight_mg_stick = st.number_input(
            "Clove Weight (mg/stick)",
            value=float(excel_data.get("clove_weight_mg_stick", 240.0)) if excel_data.get("clove_weight_mg_stick") not in (None, "") else 240.0,
            step=1.0,
        )
        stick_per_mc = st.number_input(
            "Stick per MC",
            value=int(excel_data.get("stick_per_mc", 10000)) if excel_data.get("stick_per_mc") not in (None, "") else 10000,
            step=100,
        )
        prepared_by = st.text_input("Prepared By", value=excel_data.get("prepared_by", "Lyla Isro"))

    with col2:
        standard_control = st.text_input("Standard Control", value=excel_data.get("standard_control", ""))
        flavor_standard_reference = st.text_input("Flavor Standard Reference", value=excel_data.get("flavor_standard_reference", "N/A"))
        tobacco_blend_code = st.text_input("Tobacco Blend Code", value=excel_data.get("tobacco_blend_code", "SO-18"))
        sensory_parameter = st.text_input("Sensory Parameter", value=excel_data.get("sensory_parameter", ""))
        formulation_code = st.text_input("Formulation Code", value=excel_data.get("formulation_code", ""))
        impact = st.text_input("Impact", value=excel_data.get("impact", ""))

    with col3:
        single_capsule = st.text_input("Single Capsule", value=excel_data.get("single_capsule", "N/A"))
        flavor_aroma = st.text_input("Flavor Aroma", value=excel_data.get("flavor_aroma", ""))
        double_capsule_tobacco_end = st.text_input("Double Capsule (Tobacco End)", value=excel_data.get("double_capsule_tobacco_end", ""))
        double_capsule_mouth_end = st.text_input("Double Capsule (Mouth End)", value=excel_data.get("double_capsule_mouth_end", ""))
        irritation = st.text_input("Irritation", value=excel_data.get("irritation", ""))
        cooling = st.text_input("Cooling", value=excel_data.get("cooling", ""))
        formulation_date = st.date_input("Date", value=_coerce_date_value(excel_data.get("date")))

uploaded_phase_names = set(excel_data.get("phase_metadata", {}).keys())
uploaded_phase_names.update(row.get("phase") for row in excel_materials if row.get("phase"))
uploaded_flavor_count = 2
uploaded_casing_count = 1
for phase_name in uploaded_phase_names:
    if not isinstance(phase_name, str):
        continue
    if phase_name.startswith("Flavor Pre-Mix"):
        suffix = phase_name.replace("Flavor Pre-Mix", "").strip()
        uploaded_flavor_count = max(uploaded_flavor_count, int(suffix) if suffix.isdigit() else 1)
    if phase_name.startswith("Casing Pre-Mix"):
        suffix = phase_name.replace("Casing Pre-Mix", "").strip()
        uploaded_casing_count = max(uploaded_casing_count, int(suffix) if suffix.isdigit() else 1)

with st.expander("Premix Setup", expanded=True):
    col_casing, col_flavor = st.columns(2)
    with col_casing:
        casing_premix_count = st.number_input(
            "Jumlah Casing Pre-Mix",
            min_value=1,
            max_value=20,
            value=uploaded_casing_count,
            step=1,
        )
    with col_flavor:
        flavor_premix_count = st.number_input(
            "Jumlah Flavor Pre-Mix",
            min_value=0,
            max_value=30,
            value=uploaded_flavor_count,
            step=1,
        )
    st.caption("Phase utama wajib: Casing Rajangan, Casing Krosok, Top Flavor. Premix optional dan bisa ditambah sesuai kebutuhan.")

phase_options = build_phase_options(casing_premix_count, flavor_premix_count)
if "saved_phase_metadata" not in st.session_state:
    st.session_state["saved_phase_metadata"] = excel_data.get("phase_metadata", {}).copy()

with st.expander("Phase Metadata", expanded=False):
    st.write("Isi metadata phase jika diperlukan. Semua phase dapat menambah baris material otomatis mengikuti jumlah input.")
    phase_metadata = {}
    for phase in phase_options:
        phase_values = st.session_state["saved_phase_metadata"].get(phase, excel_data.get("phase_metadata", {}).get(phase, {}))
        with st.container():
            st.markdown(f"**{phase}**")
            nav_code = st.text_input(f"{phase} NAV Item Code", value=phase_values.get("nav_code", ""), key=f"nav_{phase}")
            description = st.text_input(f"{phase} Description", value=phase_values.get("description", ""), key=f"desc_{phase}")
            blend_ratio = st.number_input(
                f"{phase} Blend Ratio",
                min_value=0.0,
                max_value=100.0,
                value=float(phase_values.get("blend_ratio", 0.0)) if phase_values.get("blend_ratio") is not None else 0.0,
                step=0.01,
                format="%.2f",
                key=f"blend_{phase}",
            )
            application = st.number_input(
                f"{phase} Application %",
                min_value=0.0,
                max_value=100.0,
                value=float(phase_values.get("application", 0.0)) if phase_values.get("application") is not None else 0.0,
                step=0.01,
                format="%.2f",
                key=f"app_{phase}",
            )
            phase_metadata[phase] = {
                "nav_code": nav_code,
                "description": description,
                "blend_ratio": blend_ratio if blend_ratio > 0 else None,
                "application": application if application > 0 else None,
            }
            if st.button(f"Save {phase}", key=f"save_phase_{phase}"):
                st.session_state["saved_phase_metadata"][phase] = phase_metadata[phase]
                st.success(f"{phase} metadata tersimpan.")

    if st.button("Save All Phase Metadata", type="secondary"):
        st.session_state["saved_phase_metadata"].update(phase_metadata)
        st.success("Semua phase metadata tersimpan.")

material_columns = [
    "phase",
    "item_code",
    "item_name",
    "dosage_input_mode",
    "dosage_mg_stick",
    "ratio_percent",
    "addition_sequence",
    "temperature",
    "agitation_rate",
    "mixing_duration",
    "work_instruction_override",
    "process_role",
    "notes",
]

default_materials = [
    {
        "phase": "Casing Rajangan",
        "item_code": "GEN-WA-00001",
        "item_name": "WATER",
        "dosage_input_mode": "mg/stick",
        "dosage_mg_stick": 1.25,
        "addition_sequence": 1,
    },
    {
        "phase": "Casing Rajangan",
        "item_code": "GEN-CHM-00015",
        "item_name": "INVERT SUGAR",
        "dosage_input_mode": "mg/stick",
        "dosage_mg_stick": 2.0,
        "addition_sequence": 2,
    },
    {
        "phase": "Casing Rajangan",
        "item_code": "GEN-CHM-00016",
        "item_name": "SV-PROPYLENE GLYCOL",
        "dosage_input_mode": "mg/stick",
        "dosage_mg_stick": 2.0,
        "addition_sequence": 2,
    },
    {
        "phase": "Casing Krosok",
        "item_code": "GEN-WA-00001",
        "item_name": "WATER",
        "dosage_input_mode": "mg/stick",
        "dosage_mg_stick": 1.0,
        "addition_sequence": 1,
    },
    {
        "phase": "Top Flavor",
        "item_code": "GEN-WA-00001",
        "item_name": "WATER",
        "dosage_input_mode": "mg/stick",
        "dosage_mg_stick": 1.0,
        "addition_sequence": 1,
    },
]

materials_df = pd.DataFrame(default_materials, columns=material_columns)
if excel_materials:
    materials_df = pd.DataFrame(excel_materials, columns=material_columns)
if "Required Fields" not in materials_df.columns:
    materials_df["Required Fields"] = ""
materials_df["Required Fields"] = materials_df.apply(
    lambda row: (
        "Dosage mg/stick"
        if str(row.get("dosage_input_mode") or "mg/stick").strip() == "mg/stick"
        else "Ratio % + Phase Application %"
    ),
    axis=1,
)
materials_df = st.data_editor(
    materials_df,
    num_rows="dynamic",
    width="stretch",
    key="material_table",
    column_config={
        "phase": st.column_config.SelectboxColumn("Phase", options=phase_options, required=True),
        "dosage_input_mode": st.column_config.SelectboxColumn("Input Mode", options=["mg/stick", "%"], required=True),
        "dosage_mg_stick": st.column_config.NumberColumn("Dosage (mg/stick)", min_value=0.0, step=0.00001, format="%.5f"),
        "ratio_percent": st.column_config.NumberColumn("Ratio %", min_value=0.0, step=0.00001, format="%.5f"),
        "addition_sequence": st.column_config.NumberColumn("Addition Sequence", min_value=0, step=1, format="%d"),
        "Required Fields": st.column_config.TextColumn("Required Fields", disabled=True),
    },
)

current_form_data = {
    "product_name": product_name,
    "formula_code": formula_code,
    "product_weight_mg_stick": product_weight_mg_stick,
    "clove_weight_mg_stick": clove_weight_mg_stick,
    "stick_per_mc": stick_per_mc,
    "prepared_by": prepared_by,
    "date": formulation_date.isoformat(),
    "standard_control": standard_control,
    "flavor_standard_reference": flavor_standard_reference,
    "tobacco_blend_code": tobacco_blend_code,
    "sensory_parameter": sensory_parameter,
    "formulation_code": formulation_code,
    "impact": impact,
    "single_capsule": single_capsule,
    "flavor_aroma": flavor_aroma,
    "double_capsule_tobacco_end": double_capsule_tobacco_end,
    "irritation": irritation,
    "double_capsule_mouth_end": double_capsule_mouth_end,
    "cooling": cooling,
    "phase_metadata": phase_metadata,
}


def collect_current_issues() -> List[Dict[str, Any]]:
    material_records = materials_df[material_columns].to_dict(orient="records")
    return (
        collect_header_issues(
            product_name,
            formula_code,
            prepared_by,
            product_weight_mg_stick,
            clove_weight_mg_stick,
            stick_per_mc,
        )
        + collect_material_issues(material_records, phase_metadata)
    )


save_col, generate_col = st.columns([1, 2])
with save_col:
    save_clicked = st.button("Save Data / Validate Draft", type="secondary")
with generate_col:
    generate_clicked = st.button("Generate Formulation Sheet", type="primary")

if save_clicked:
    st.session_state["saved_form_data"] = current_form_data
    st.session_state["saved_materials"] = materials_df[material_columns].to_dict(orient="records")
    st.session_state["validation_issues"] = collect_current_issues()

if "validation_issues" in st.session_state:
    issues = st.session_state["validation_issues"]
    if issues:
        st.warning(f"{len(issues)} data perlu dilengkapi sebelum generate.")
        issues_df = pd.DataFrame(issues)
        st.dataframe(issues_df, width="stretch", hide_index=True)
        material_issue_rows = sorted(
            {
                issue["Row"]
                for issue in issues
                if isinstance(issue.get("Row"), int)
            }
        )
        if material_issue_rows:
            with st.expander("Preview row material yang perlu dicek", expanded=True):
                preview_df = materials_df.iloc[[row - 1 for row in material_issue_rows if row - 1 < len(materials_df)]]
                st.dataframe(preview_df, width="stretch")
    else:
        st.success("Draft tersimpan dan valid. Data siap digenerate.")

model_formulation_input = build_formulation_input_from_dict(
    current_form_data,
    materials_df[material_columns].to_dict(orient="records"),
)
model_json_bytes = export_formulation_model_json(model_formulation_input)
model_csv_bytes = export_formulation_model_csv_bundle(model_formulation_input)
export_col_json, export_col_csv = st.columns(2)
with export_col_json:
    st.download_button(
        "Download Model JSON",
        model_json_bytes,
        file_name=f"Formulation_Model_{formula_code.strip() or 'model'}.json",
        mime="application/json",
    )
with export_col_csv:
    st.download_button(
        "Download Model CSV Bundle",
        model_csv_bytes,
        file_name=f"Formulation_Model_{formula_code.strip() or 'model'}.zip",
        mime="application/zip",
    )

if generate_clicked:
    if not TEMPLATE_PATH.exists():
        st.error(f"Template tidak ditemukan: {TEMPLATE_PATH}")
    else:
        try:
            preflight_issues = collect_current_issues()
            st.session_state["validation_issues"] = preflight_issues
            if preflight_issues:
                st.error("Generate dibatalkan. Perbaiki data berikut terlebih dahulu:")
                st.dataframe(pd.DataFrame(preflight_issues), width="stretch", hide_index=True)
                st.stop()

            material_records = materials_df[material_columns].to_dict(orient="records")
            formulation_input = build_formulation_input_from_dict(current_form_data, material_records)
            safe_formula_code = "".join(char if char.isalnum() or char in ("-", "_") else "_" for char in formula_code.strip()) or "Formulation"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = OUTPUT_DIR / f"Generated_Formulation_Sheet_{safe_formula_code}_{timestamp}.xlsx"
            result_path = generate_formulation_workbook(formulation_input, TEMPLATE_PATH, output_path)
            st.success(f"Formulasi berhasil dibuat: {result_path.name}")
            with open(result_path, "rb") as f:
                st.download_button(
                    "Download hasil formulasi",
                    f,
                    file_name=result_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as exc:
            st.error(str(exc))
