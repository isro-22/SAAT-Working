import unittest
import json
import csv
import io
import zipfile
from copy import deepcopy
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

from formulation_generator import (
    FORMULA_MATERIAL_COLUMNS,
    SECTION_ROW_RANGES,
    build_formulation_code_formula,
    build_formulation_input_from_dict,
    create_blank_input_workbook,
    export_formulation_model_json,
    export_formulation_model_csv_bundle,
    generate_formulation_workbook,
)


class TestFormulationGenerator(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.root_dir = Path(__file__).resolve().parents[1]
        cls.template_path = cls.root_dir / "templates" / "Template_Generate.xlsm"
        if not cls.template_path.exists():
            raise FileNotFoundError(f"Test template not found: {cls.template_path}")

        cls.form_data = {
            "product_name": "SAAT 369",
            "formula_code": "CS_K-CK-LCDT-04_R00",
            "product_weight_mg_stick": 720,
            "clove_weight_mg_stick": 240,
            "stick_per_mc": 10000,
            "prepared_by": "Lyla Isro",
            "date": "2026-07-07",
            "standard_control": "DJARUM 76",
            "flavor_standard_reference": "N/A",
            "tobacco_blend_code": "SO-18",
            "sensory_parameter": "Sample",
            "formulation_code": "CS_K-CK-LCDT-04_R00",
            "impact": "Low",
            "single_capsule": "N/A",
            "flavor_aroma": "Spice",
            "double_capsule_tobacco_end": "N/A",
            "irritation": "None",
            "double_capsule_mouth_end": "N/A",
            "cooling": "Mild",
            "phase_metadata": {
                "Casing Rajangan": {
                    "nav_code": "GEN-CS-00000",
                    "description": "CS_R-CK-LCDT-04_R00",
                    "blend_ratio": 0.38,
                    "application": 0.03078,
                },
            },
        }
        cls.material_rows = [
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.25,
                "addition_sequence": 1,
            },
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 2,
            },
            {
                "phase": "Casing Krosok",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
        ]

    def build_formulation(self, form_data=None, material_rows=None):
        return build_formulation_input_from_dict(
            deepcopy(form_data if form_data is not None else self.form_data),
            deepcopy(material_rows if material_rows is not None else self.material_rows),
        )

    def with_required_main_phases(self, material_rows):
        rows = deepcopy(material_rows)
        existing_phases = {row.get("phase") for row in rows}
        required_defaults = {
            "Casing Rajangan": {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            "Casing Krosok": {
                "phase": "Casing Krosok",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            "Top Flavor": {
                "phase": "Top Flavor",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
        }
        for phase, row in required_defaults.items():
            if phase not in existing_phases:
                rows.append(row)
        return rows

    def generate_workbook(self, form_data=None, material_rows=None):
        formulation = self.build_formulation(form_data, material_rows)
        temp_dir = TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        output_path = Path(temp_dir.name) / "formulation_test.xlsx"
        generate_formulation_workbook(formulation, self.template_path, output_path)
        return openpyxl.load_workbook(output_path, data_only=False)

    def assert_formula_or_value(self, actual, expected, formula_prefix=None):
        if isinstance(actual, str) and actual.startswith("="):
            if formula_prefix is not None:
                self.assertTrue(
                    actual.startswith(formula_prefix),
                    f"Expected formula beginning with {formula_prefix!r}, got {actual!r}",
                )
            return
        self.assertAlmostEqual(actual, expected)

    def test_generate_formulation_workbook_creates_xlsx(self):
        formulation = self.build_formulation()
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "generated" / "formulation_test.xlsx"
            result_path = generate_formulation_workbook(formulation, self.template_path, output_path)

            self.assertTrue(result_path.exists())
            self.assertEqual(result_path, output_path)

            workbook = openpyxl.load_workbook(result_path)
            self.assertIn("INPUT", workbook.sheetnames)
            self.assertIn("Formula", workbook.sheetnames)

            with zipfile.ZipFile(result_path) as workbook_zip:
                self.assertFalse(
                    any(name.startswith("xl/externalLinks/") for name in workbook_zip.namelist()),
                    "Generated .xlsx should not preserve external links from the source template.",
                )
                self.assertTrue(
                    any(name.startswith("xl/media/") for name in workbook_zip.namelist()),
                    "Generated .xlsx should include the project logo image.",
                )

    def test_export_formulation_model_json_has_database_like_structure(self):
        formulation = self.build_formulation()
        payload = json.loads(export_formulation_model_json(formulation).decode("utf-8"))

        self.assertEqual(payload["schema_version"], "1.0")
        self.assertEqual(payload["model_type"], "formulation_product")
        self.assertIn("product", payload)
        self.assertIn("phase_metadata", payload)
        self.assertIn("materials", payload)
        self.assertEqual(payload["product"]["formula_code"], "CS_K-CK-LCDT-04_R00")
        self.assertIsInstance(payload["phase_metadata"], list)
        self.assertIsInstance(payload["materials"], list)
        self.assertGreaterEqual(len(payload["materials"]), 1)
        self.assertEqual(payload["materials"][0]["phase"], "Casing Rajangan")

    def test_export_formulation_model_csv_bundle_contains_three_tables(self):
        formulation = self.build_formulation()
        bundle = export_formulation_model_csv_bundle(formulation)

        with zipfile.ZipFile(io.BytesIO(bundle)) as archive:
            self.assertEqual(
                sorted(archive.namelist()),
                ["materials.csv", "phase_metadata.csv", "product.csv"],
            )
            with archive.open("product.csv") as handle:
                product_rows = list(csv.DictReader(io.TextIOWrapper(handle, encoding="utf-8")))
            with archive.open("phase_metadata.csv") as handle:
                phase_rows = list(csv.DictReader(io.TextIOWrapper(handle, encoding="utf-8")))
            with archive.open("materials.csv") as handle:
                material_rows = list(csv.DictReader(io.TextIOWrapper(handle, encoding="utf-8")))

        self.assertEqual(product_rows[0]["formula_code"], "CS_K-CK-LCDT-04_R00")
        self.assertGreaterEqual(len(phase_rows), 1)
        self.assertGreaterEqual(len(material_rows), 1)
        self.assertEqual(material_rows[0]["phase"], "Casing Rajangan")

    def test_generated_workbook_contains_metadata(self):
        formulation = self.build_formulation()
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "formulation_test.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path)

            formula_sheet = workbook["Formula"]
            self.assertIn("A1:B3", [str(cell_range) for cell_range in formula_sheet.merged_cells.ranges])
            self.assertIsNone(formula_sheet["A1"].value)
            self.assertEqual(formula_sheet.cell(row=5, column=3).value, "SAAT 369")
            self.assertEqual(formula_sheet.cell(row=6, column=1).value, "Mixing Factory")
            self.assertEqual(formula_sheet.cell(row=6, column=3).value, "PT SAAT")
            self.assertEqual(formula_sheet.cell(row=7, column=3).value, "Lyla Isro")
            self.assertEqual(formula_sheet.cell(row=1, column=16).value, "Document No")
            self.assertEqual(formula_sheet.cell(row=1, column=17).value, "BOL-ID-BOLL-SPEC-003")
            self.assertEqual(formula_sheet.cell(row=2, column=16).value, "Revision No")
            self.assertEqual(formula_sheet.cell(row=2, column=17).value, "00")
            self.assertEqual(formula_sheet.cell(row=3, column=16).value, "Effective Date")
            self.assertEqual(formula_sheet.cell(row=3, column=17).value, "2026-07-07")
            self.assertEqual(formula_sheet.cell(row=10, column=8).value, "Flavor Standard Reference")
            self.assertEqual(formula_sheet.cell(row=10, column=10).value, "N/A")
            self.assertEqual(formula_sheet.cell(row=12, column=8).value, "Sensory Parameter")
            self.assertEqual(formula_sheet.cell(row=13, column=8).value, "Impact")
            self.assertEqual(formula_sheet.cell(row=14, column=8).value, "Flavor Aroma")
            self.assertEqual(formula_sheet.cell(row=15, column=8).value, "Irritation")
            self.assertEqual(formula_sheet.cell(row=16, column=8).value, "Cooling")
            self.assertEqual(formula_sheet.cell(row=21, column=5).value, "GEN-CS-00000")
            self.assertEqual(formula_sheet.cell(row=22, column=5).value, "CS_R-CK-LCDT-04_R00")
            self.assertEqual(formula_sheet.cell(row=13, column=3).value, build_formulation_code_formula())

    def test_casing_top_flavor_summary_uses_phase_description_totals_and_usd_price(self):
        workbook = self.generate_workbook()
        formula_sheet = workbook["Formula"]

        header_row = None
        for row in range(1, formula_sheet.max_row + 1):
            value = formula_sheet.cell(row=row, column=2).value
            if isinstance(value, str) and "Casing & Top Flavor" in value:
                header_row = row
                break

        self.assertIsNotNone(header_row)
        first_summary_row = header_row + 1
        first_total_row = SECTION_ROW_RANGES["Casing Rajangan"][-1] + 1

        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=2).value, "=E22")
        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=3).value, f"=H{first_total_row}")
        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=4).value, "=E23")
        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=7).value, f"=J{first_total_row}")
        self.assertEqual(
            formula_sheet.cell(row=first_summary_row, column=8).value,
            f"=G{first_summary_row}*(C{first_summary_row}/1000000)*1000",
        )
        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=7).number_format, '"$"#,##0.00')
        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=8).number_format, '"$"#,##0.00')

    def test_approval_block_contains_prepared_reviewed_and_approved_columns(self):
        workbook = self.generate_workbook()
        formula_sheet = workbook["Formula"]

        approval_row = None
        for row in range(1, formula_sheet.max_row + 1):
            if formula_sheet.cell(row=row, column=1).value == "Prepared By":
                approval_row = row
                break

        self.assertIsNotNone(approval_row)
        self.assertEqual(formula_sheet.cell(row=approval_row, column=1).value, "Prepared By")
        self.assertEqual(formula_sheet.cell(row=approval_row, column=6).value, "Reviewed By")
        self.assertEqual(formula_sheet.cell(row=approval_row, column=13).value, "Approved By")
        self.assertEqual(formula_sheet.cell(row=approval_row + 2, column=1).value, "Name : Lyla Isro")
        self.assertEqual(formula_sheet.cell(row=approval_row + 2, column=6).value, "Name : Mochamad Setyawan")
        self.assertEqual(formula_sheet.cell(row=approval_row + 2, column=13).value, "Name : Andrew Yip")

    def test_blank_input_workbook_can_be_downloaded_as_xlsx(self):
        workbook_bytes = create_blank_input_workbook(self.template_path)
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "blank_input.xlsx"
            output_path.write_bytes(workbook_bytes)
            workbook = openpyxl.load_workbook(output_path)

        self.assertEqual(workbook.sheetnames, ["INPUT"])
        input_sheet = workbook["INPUT"]
        self.assertIsNone(input_sheet.cell(row=2, column=2).value)
        self.assertEqual(input_sheet.cell(row=4, column=4).value, "Casing Rajangan")
        self.assertIsNone(input_sheet.cell(row=4, column=5).value)
        self.assertIsNone(input_sheet.cell(row=12, column=1).value)

    def test_material_section_has_written_rows(self):
        formulation = self.build_formulation()
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "formulation_test.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path)

            formula_sheet = workbook["Formula"]
            first_row = formula_sheet.cell(row=26, column=2).value
            second_row = formula_sheet.cell(row=27, column=2).value
            self.assertEqual(first_row, "GEN-WA-00001")
            self.assertEqual(second_row, "GEN-CHM-00015")

    def test_material_sections_resize_to_written_materials_and_totals_match(self):
        workbook = self.generate_workbook()
        formula_sheet = workbook["Formula"]

        row_indices = SECTION_ROW_RANGES["Casing Rajangan"]
        self.assertEqual(len(row_indices), 2)
        total_row = row_indices[-1] + 1
        self.assert_formula_or_value(formula_sheet.cell(row=total_row, column=7).value, 1.0, "=SUM(G26:")
        self.assert_formula_or_value(formula_sheet.cell(row=total_row, column=8).value, 3.25, "=SUM(H26:")
        self.assert_formula_or_value(formula_sheet.cell(row=total_row, column=10).value, 0.8492307692307692, "=SUM(J26:")
        self.assert_formula_or_value(formula_sheet.cell(row=total_row, column=11).value, 0.0325, "=SUM(K26:")

    def test_material_calculations_include_ratio_formulation_price_and_dosage_kg_mc(self):
        workbook = self.generate_workbook()
        formula_sheet = workbook["Formula"]

        total_dosage = 3.25
        expected_rows = {
            26: {"ratio": 1.25 / total_dosage, "dosage_kg_mc": 0.0125},
            27: {"ratio": 2.0 / total_dosage, "dosage_kg_mc": 0.02},
        }

        for row, expected in expected_rows.items():
            material_price = formula_sheet.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["material_price"]).value or 0
            self.assertAlmostEqual(
                formula_sheet.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["ratio"]).value,
                expected["ratio"],
            )
            self.assertAlmostEqual(
                formula_sheet.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["formulation_price"]).value,
                expected["ratio"] * material_price,
            )
            self.assertAlmostEqual(
                formula_sheet.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["dosage_kg_mc"]).value,
                expected["dosage_kg_mc"],
            )

    def test_percent_mode_converts_to_dosage_mg_stick_with_five_decimals(self):
        material_rows = self.with_required_main_phases([
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_input_mode": "%",
                "ratio_percent": 38,
                "application_percent": 3.078,
                "addition_sequence": 1,
            },
        ])
        workbook = self.generate_workbook(material_rows=material_rows)
        formula_sheet = workbook["Formula"]

        expected_dosage = round(0.38 * 0.03078 * (720 - 240), 5)
        self.assertEqual(formula_sheet.cell(row=26, column=FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"]).value, expected_dosage)
        self.assertEqual(formula_sheet.cell(row=26, column=FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"]).number_format, "0.00000")

    def test_percent_mode_uses_phase_metadata_application_without_material_application_column(self):
        form_data = deepcopy(self.form_data)
        form_data["phase_metadata"]["Top Flavor"] = {
            "nav_code": "TF-GA-09",
            "description": "TF-GA-09",
            "blend_ratio": 100,
            "application": 2.14,
        }
        material_rows = [
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Casing Krosok",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_input_mode": "%",
                "ratio_percent": 20,
                "addition_sequence": 1,
            },
        ]

        workbook = self.generate_workbook(form_data=form_data, material_rows=material_rows)
        formula_sheet = workbook["Formula"]

        expected_dosage = round(0.20 * 0.0214 * (720 - 240), 5)
        self.assertEqual(formula_sheet.cell(row=52, column=FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"]).value, expected_dosage)

    def test_formula_output_is_sorted_by_addition_sequence_within_phase(self):
        material_rows = [
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 20,
            },
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.25,
                "addition_sequence": 10,
            },
        ]
        workbook = self.generate_workbook(material_rows=self.with_required_main_phases(material_rows))
        formula_sheet = workbook["Formula"]

        self.assertEqual(formula_sheet.cell(row=26, column=2).value, "GEN-WA-00001")
        self.assertEqual(formula_sheet.cell(row=26, column=13).value, 10)
        self.assertEqual(formula_sheet.cell(row=27, column=2).value, "GEN-CHM-00015")
        self.assertEqual(formula_sheet.cell(row=27, column=13).value, 20)

    def test_process_instruction_columns_merge_by_addition_sequence(self):
        material_rows = [
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
                "temperature": "Ambient",
                "agitation_rate": "500 RPM",
                "mixing_duration": "5 min",
                "work_instruction_override": "Start mixing.",
            },
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 2,
            },
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-CHM-00016",
                "item_name": "SV-PROPYLENE GLYCOL",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 2,
                "temperature": "Ambient",
                "agitation_rate": "500 RPM",
                "mixing_duration": "15 min",
                "work_instruction_override": "Mix till all solids are dissolved.",
            },
            {
                "phase": "Casing Krosok",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
        ]
        workbook = self.generate_workbook(material_rows=material_rows)
        formula_sheet = workbook["Formula"]
        merged_ranges = {str(cell_range) for cell_range in formula_sheet.merged_cells.ranges}

        for column_name in ("addition_sequence", "temperature", "agitation_rate", "mixing_duration", "work_instruction"):
            column = FORMULA_MATERIAL_COLUMNS[column_name]
            column_letter = openpyxl.utils.get_column_letter(column)
            self.assertIn(f"{column_letter}27:{column_letter}28", merged_ranges)

        self.assertEqual(formula_sheet.cell(row=27, column=FORMULA_MATERIAL_COLUMNS["addition_sequence"]).value, 2)
        self.assertEqual(formula_sheet.cell(row=27, column=FORMULA_MATERIAL_COLUMNS["temperature"]).value, "Ambient")
        self.assertEqual(formula_sheet.cell(row=27, column=FORMULA_MATERIAL_COLUMNS["agitation_rate"]).value, "500 RPM")
        self.assertEqual(formula_sheet.cell(row=27, column=FORMULA_MATERIAL_COLUMNS["mixing_duration"]).value, "15 min")
        self.assertEqual(
            formula_sheet.cell(row=27, column=FORMULA_MATERIAL_COLUMNS["work_instruction"]).value,
            "Mix till all solids are dissolved.",
        )

    def test_invalid_phase_dosage_and_stick_per_mc_are_rejected(self):
        cases = [
            ("unsupported phase", self.form_data, [{**self.material_rows[0], "phase": "Unknown Phase"}]),
            ("zero dosage", self.form_data, [{**self.material_rows[0], "dosage_mg_stick": 0}]),
            ("negative dosage", self.form_data, [{**self.material_rows[0], "dosage_mg_stick": -1}]),
            ("zero stick_per_mc", {**self.form_data, "stick_per_mc": 0}, self.material_rows),
            ("negative stick_per_mc", {**self.form_data, "stick_per_mc": -100}, self.material_rows),
        ]

        for name, form_data, material_rows in cases:
            with self.subTest(name=name):
                with TemporaryDirectory() as temp_dir:
                    output_path = Path(temp_dir) / "invalid.xlsx"
                    with self.assertRaises(ValueError):
                        formulation = self.build_formulation(form_data, material_rows)
                        generate_formulation_workbook(formulation, self.template_path, output_path)

    def test_top_flavor_phase_expands_when_materials_exceed_base_capacity(self):
        top_flavor_rows = [
            {
                "phase": "Top Flavor",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 2,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-CHM-00016",
                "item_name": "SV-PROPYLENE GLYCOL",
                "dosage_mg_stick": 3.0,
                "addition_sequence": 3,
            },
        ]
        formulation = self.build_formulation(material_rows=self.with_required_main_phases(top_flavor_rows))

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "top_flavor_expanded.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            formula_sheet = workbook["Formula"]
            top_flavor_rows_written = SECTION_ROW_RANGES["Top Flavor"][: len(top_flavor_rows)]

            self.assertGreaterEqual(len(SECTION_ROW_RANGES["Top Flavor"]), len(top_flavor_rows))
            self.assertEqual(
                [formula_sheet.cell(row=row, column=2).value for row in top_flavor_rows_written],
                ["GEN-WA-00001", "GEN-CHM-00015", "GEN-CHM-00016"],
            )
            self.assertEqual(formula_sheet.cell(row=13, column=3).value, build_formulation_code_formula())

    def test_premix_phase_expands_when_materials_exceed_base_capacity(self):
        premix_rows = []
        for index in range(12):
            premix_rows.append({
                "phase": "Casing Pre-Mix",
                "item_code": "GEN-WA-00001" if index % 2 == 0 else "GEN-CHM-00015",
                "item_name": "WATER" if index % 2 == 0 else "INVERT SUGAR",
                "dosage_mg_stick": 1.0 + index,
                "addition_sequence": index + 1,
            })
        formulation = self.build_formulation(material_rows=self.with_required_main_phases(premix_rows))

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "premix_over_capacity.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            formula_sheet = workbook["Formula"]
            premix_rows_written = SECTION_ROW_RANGES["Casing Pre-Mix"][: len(premix_rows)]
            self.assertGreaterEqual(len(SECTION_ROW_RANGES["Casing Pre-Mix"]), len(premix_rows))
            self.assertEqual(
                [formula_sheet.cell(row=row, column=2).value for row in premix_rows_written[:3]],
                ["GEN-WA-00001", "GEN-CHM-00015", "GEN-WA-00001"],
            )

    def test_additional_flavor_premix_phase_creates_dynamic_section(self):
        premix_rows = [
            {
                "phase": "Flavor Pre-Mix 3",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Flavor Pre-Mix 3",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 2,
            },
        ]
        formulation = self.build_formulation(material_rows=self.with_required_main_phases(premix_rows))

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "flavor_premix_3.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            formula_sheet = workbook["Formula"]

            self.assertIn("Flavor Pre-Mix 3", SECTION_ROW_RANGES)
            first_row = SECTION_ROW_RANGES["Flavor Pre-Mix 3"][0]
            self.assertEqual(formula_sheet.cell(row=first_row, column=2).value, "GEN-WA-00001")
            self.assertEqual(formula_sheet.cell(row=first_row + 1, column=2).value, "GEN-CHM-00015")


if __name__ == "__main__":
    unittest.main()
