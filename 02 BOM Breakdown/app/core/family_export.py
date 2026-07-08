from __future__ import annotations

import re
from io import BytesIO

import pandas as pd
from openpyxl.styles import PatternFill

from .excel_io import load_excel
from .family_compare import build_family_matrix
from .vba_compat import clean_text


def qty_columns(frame: pd.DataFrame) -> list[str]:
    return [col for col in frame.columns if col.endswith(" Qty")]


def changed_qty_mask(frame: pd.DataFrame) -> pd.DataFrame:
    mask = pd.DataFrame(False, index=frame.index, columns=frame.columns)
    qty_cols = qty_columns(frame)
    if not qty_cols:
        return mask
    for idx, row in frame[qty_cols].iterrows():
        values = {clean_text(value) for value in row.tolist() if clean_text(value)}
        if len(values) > 1:
            mask.loc[idx, qty_cols] = True
    return mask


def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", clean_text(name) or "Family")[:31]
    candidate = cleaned
    counter = 2
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def export_family_workbook(file_bytes: bytes, include_all_versions: bool = True) -> bytes:
    export_data = load_excel(file_bytes)
    families = sorted(x for x in export_data.cf["Family Name"].map(clean_text).unique() if x)
    output = BytesIO()
    yellow = PatternFill(fill_type="solid", fgColor="FFF3B0")
    used_sheet_names: set[str] = set()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_rows = []
        for family_name in families:
            matrix = build_family_matrix(export_data, family_name, include_all_versions=include_all_versions)
            if matrix.empty:
                continue
            sheet_name = safe_sheet_name(family_name, used_sheet_names)
            matrix.to_excel(writer, sheet_name=sheet_name, index=False)
            mask = changed_qty_mask(matrix)
            worksheet = writer.sheets[sheet_name]
            for row_idx, row in enumerate(mask.itertuples(index=False), start=2):
                for col_idx, changed in enumerate(row, start=1):
                    if changed:
                        worksheet.cell(row=row_idx, column=col_idx).fill = yellow
            worksheet.freeze_panes = "A2"
            summary_rows.append({"Family Name": family_name, "Rows": len(matrix), "Sheet": sheet_name})
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)

    return output.getvalue()
